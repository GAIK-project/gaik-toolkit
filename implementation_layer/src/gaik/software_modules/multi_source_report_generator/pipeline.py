"""Generic multi-source report generator.

Turns a mixed set of source files (PDF, Word, Excel/CSV, text, Markdown,
audio/video, images) into a long-form Markdown report whose structure is
defined entirely by the user (section titles + per-section instructions).

Design (see the module README):
  * Lean GAIK-module constructor: ``__init__(*, api_config=None, use_azure=True)``.
  * One public method: ``run(...)`` — all per-run settings live here.
  * Pure orchestration over existing GAIK components. Nothing in this module is
    specific to any report domain (no classification, relevance scoring, file
    selection, or fixed section templates).

Pipeline:
    input_paths -> evidence (parse/transcribe/extract each file to markdown)
                -> ONE LLM call that writes the whole report (all user-defined
                   sections + optional sample report as a format/style template)
                -> Markdown report (+ per-section breakdown split from it)
"""

from __future__ import annotations

import csv
import json
import logging
import re
from collections.abc import Callable
from pathlib import Path
from typing import Any

from .models import (
    EvidenceItem,
    GeneratedSection,
    ReportGenerationResult,
    ReportSectionSpec,
)

# LLM path (monkeypatchable in tests). Imported defensively so the module can be
# imported even in environments without the openai SDK present.
try:
    from gaik.software_components.llm import create_llm_client
except Exception:  # pragma: no cover - optional at import time
    create_llm_client = None  # type: ignore

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Supported source types (by file extension)
# ---------------------------------------------------------------------------

_TEXT_EXT = {".txt"}
_MARKDOWN_EXT = {".md", ".markdown"}
_PDF_EXT = {".pdf"}
_DOCX_EXT = {".docx"}
_CSV_EXT = {".csv"}
_XLSX_EXT = {".xlsx", ".xls"}
_AUDIO_EXT = {".mp3", ".wav", ".m4a", ".aac", ".flac", ".ogg"}
_VIDEO_EXT = {".mp4", ".mov", ".mkv", ".avi", ".webm"}
_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".webp", ".tiff", ".tif", ".bmp", ".gif"}

_EXT_TO_TYPE: dict[str, str] = {
    **{e: "text" for e in _TEXT_EXT},
    **{e: "markdown" for e in _MARKDOWN_EXT},
    **{e: "pdf" for e in _PDF_EXT},
    **{e: "docx" for e in _DOCX_EXT},
    **{e: "csv" for e in _CSV_EXT},
    **{e: "xlsx" for e in _XLSX_EXT},
    **{e: "audio" for e in _AUDIO_EXT},
    **{e: "video" for e in _VIDEO_EXT},
    **{e: "image" for e in _IMAGE_EXT},
}

SUPPORTED_EXTENSIONS = frozenset(_EXT_TO_TYPE)


# ---------------------------------------------------------------------------
# Prompt templates (generic — no domain-specific wording)
# ---------------------------------------------------------------------------

_SYSTEM_PROMPT = """\
You are a professional report writer. You write a COMPLETE report from the \
evidence the user provides.

VERY IMPORTANT:
- Do not make up anything. Generate content only from the given evidence. If the \
evidence does not contain information for a section, state this clearly within \
that section rather than inventing or inferring content.
- The EVIDENCE is your ONLY source of content. A FORMAT REFERENCE (if given) is a \
layout example, usually about a completely DIFFERENT subject — copy its shape, \
never its content. Never reuse its facts, topic, names, numbers, examples, or wording.
- Use simple, direct language without fluff, filler phrases, or em dashes (—).
- Maintain a neutral and professional tone throughout.

Rules:
- Use only the provided evidence. Do not invent facts, figures, names, or events.
- Output clean Markdown. Begin with a single level-1 title (`# <report title>`).
- Include EXACTLY the requested sections, in the given order, each as a level-2 \
heading (`## <section title>`) using the exact titles provided. Do not add, drop, \
merge, or reorder sections.
- Follow each section's content instructions. If the evidence lacks information \
for a section, say so explicitly within that section.
- FORMAT: If a FORMAT REFERENCE is provided, it governs BOTH structure AND length. \
Reproduce its formatting choices exactly — how each section is structured \
internally, whether it uses prose paragraphs, bullet lists, or numbered lists, \
the approximate number of paragraphs or bullet points per section, the approximate \
length of each paragraph or item, bold lead-in patterns, and citation style — but \
take ZERO content from it. The reference is about a different topic; if you find \
yourself repeating any fact, name, number, or sentence from it, stop and write from \
the evidence instead. Length rule: each section should be approximately the same \
total length and density as the corresponding reference section — not shorter and \
not longer. Do not pad to seem thorough. If NO reference is provided, write in a \
clean, professional report format of your choice.

Content guidelines:
- When a FORMAT REFERENCE is provided, match its level of brevity and detail. Do \
not write more than the reference demonstrates — "comprehensive coverage" means \
matching the reference's density, not exhausting all available evidence.
- When NO FORMAT REFERENCE is provided: make the most of all relevant information \
and cover each section fully without omitting important details.
- Use direct language that precisely communicates facts and insights.
- Vary sentence structures naturally to maintain reader engagement.
- Use domain-specific terminology where appropriate, but prefer simpler terms \
when they communicate the same meaning.
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _slug(text: str, *, max_len: int = 40) -> str:
    keep = [c.lower() if c.isalnum() else "_" for c in text.strip()]
    s = "".join(keep)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")[:max_len] or "section"


def _normalize_sections(
    sections: list[ReportSectionSpec | dict[str, Any]],
) -> list[ReportSectionSpec]:
    if not sections:
        raise ValueError("`sections` must contain at least one section.")
    out: list[ReportSectionSpec] = []
    for i, s in enumerate(sections):
        if isinstance(s, ReportSectionSpec):
            # Copy so we can fill in a derived id without mutating the caller's object.
            spec = ReportSectionSpec(
                title=s.title,
                instructions=s.instructions,
                required=s.required,
                id=s.id,
                depends_on=list(s.depends_on),
            )
        elif isinstance(s, dict):
            if not s.get("title"):
                raise ValueError(f"Section {i} is missing a 'title'.")
            spec = ReportSectionSpec(
                title=str(s["title"]),
                instructions=str(s.get("instructions", "")),
                required=bool(s.get("required", True)),
                id=str(s["id"]) if s.get("id") else None,
                depends_on=[str(d) for d in (s.get("depends_on") or [])],
            )
        else:
            raise TypeError(
                f"Section {i} must be a ReportSectionSpec or dict, got {type(s).__name__}."
            )
        # Derive a stable id from the title when not given.
        if not spec.id:
            spec.id = _slug(spec.title)
        out.append(spec)

    # Validate ids and dependencies.
    seen_ids: set[str] = set()
    for spec in out:
        if spec.id in seen_ids:
            raise ValueError(
                f"Duplicate section id '{spec.id}' (derived from title '{spec.title}'). "
                "Give the colliding sections explicit, unique `id` values."
            )
        seen_ids.add(spec.id)
    for spec in out:
        for dep in spec.depends_on:
            if dep == spec.id:
                raise ValueError(f"Section '{spec.id}' cannot depend on itself.")
            if dep not in seen_ids:
                raise ValueError(
                    f"Section '{spec.id}' depends on unknown section id '{dep}'. "
                    f"Known ids: {sorted(seen_ids)}."
                )
    return out


def _collect_input_files(input_paths: list[str | Path]) -> list[Path]:
    """Expand files/folders into a flat list of supported source files."""
    if not input_paths:
        raise ValueError("`input_paths` must contain at least one file or folder.")
    collected: list[Path] = []
    seen: set[Path] = set()
    for raw in input_paths:
        p = Path(raw)
        if not p.exists():
            raise FileNotFoundError(f"Input path does not exist: {p}")
        candidates = sorted(p.rglob("*")) if p.is_dir() else [p]
        for c in candidates:
            if not c.is_file():
                continue
            if c.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            rp = c.resolve()
            if rp in seen:
                continue
            seen.add(rp)
            collected.append(c)
    return collected


def _csv_to_markdown(path: Path) -> str:
    with open(path, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    if not rows:
        return "_(empty CSV)_"
    return _rows_to_markdown_table(rows)


def _rows_to_markdown_table(rows: list[list[Any]]) -> str:
    width = max(len(r) for r in rows)
    norm = [[str(c) if c is not None else "" for c in r] + [""] * (width - len(r)) for r in rows]
    header, *body = norm
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    lines += ["| " + " | ".join(r) + " |" for r in body]
    return "\n".join(lines)


def _xlsx_to_markdown(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning(
            "Cannot parse %s: 'openpyxl' is not installed. Install gaik extras "
            "for spreadsheet support. Skipping spreadsheet content.",
            path.name,
        )
        return f"_([.xlsx not parsed: openpyxl not installed] {path.name})_"

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    blocks: list[str] = []
    for ws in wb.worksheets:
        rows = [[("" if v is None else v) for v in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(str(c).strip() for c in r)]
        if not rows:
            continue
        blocks.append(f"### Sheet: {ws.title}\n\n{_rows_to_markdown_table(rows)}")
    wb.close()
    return "\n\n".join(blocks) if blocks else "_(no tabular data found)_"


def _normalize_report_markdown(text: str | None, report_title: str) -> str:
    """Ensure the model's report starts with a single H1 title and ends cleanly."""
    body = (text or "").strip()
    if not body:
        body = f"# {report_title}"
    elif not body.lstrip().startswith("# "):
        body = f"# {report_title}\n\n{body}"
    return body + "\n"


def _split_into_sections(markdown: str, specs: list[ReportSectionSpec]) -> list[GeneratedSection]:
    """Split the assembled report into per-section objects by its level-2 headings.

    Positional, robust to minor heading rewording: the Nth ``##`` heading maps to
    the Nth requested section title (for stable filenames), with the body taken
    verbatim from the report. Falls back to a single section if no ``##`` headings
    are present.
    """
    lines = markdown.splitlines()
    # Strip any accidental "SECTION:" prefix the model may echo from the prompt.
    heads = [
        (i, line[3:].strip().removeprefix("SECTION:").strip())
        for i, line in enumerate(lines)
        if line.startswith("## ")
    ]
    if not heads:
        return [
            GeneratedSection(
                title=specs[0].title if specs else "Report", content_markdown=markdown.strip()
            )
        ]

    sections: list[GeneratedSection] = []
    for k, (idx, heading_title) in enumerate(heads):
        end = heads[k + 1][0] if k + 1 < len(heads) else len(lines)
        body = "\n".join(lines[idx + 1 : end]).strip()
        title = specs[k].title if k < len(specs) else heading_title
        sections.append(GeneratedSection(title=title, content_markdown=body))
    return sections


# ---------------------------------------------------------------------------
# Sample-report section matching (used by the agentic path only)
# ---------------------------------------------------------------------------


def _normalize_heading(text: str) -> str:
    """Normalize a heading for tolerant matching.

    Lowercase; strip leading markdown markers and section numbering; drop
    punctuation; collapse whitespace.
    """
    t = text.strip()
    t = re.sub(r"^#+\s*", "", t)  # markdown heading markers
    t = re.sub(r"^\d+([.)]\d*)*[.)]?\s+", "", t)  # leading numbering e.g. "1." / "1.2)"
    t = t.lower()
    t = re.sub(r"[^\w\s]", " ", t)  # punctuation -> space
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _split_sample_sections(sample_markdown: str) -> dict[str, tuple[str, str]]:
    """Split a sample report at its top-level report-section headings.

    Sample reports must use ``##`` Markdown headings for their main sections.
    Detects the shallowest heading level *below* a single document title
    (typically ``##``) and splits only on that level, so each returned block
    keeps its own ``###``/``####`` subsections intact. Returns
    ``{normalized_heading: (original_heading, section_markdown)}`` where the
    section markdown includes the heading line itself.
    """
    if not sample_markdown:
        return {}
    lines = sample_markdown.splitlines()
    headings: list[tuple[int, int, str]] = []  # (line_index, level, text)
    for i, line in enumerate(lines):
        m = re.match(r"^(#{1,6})\s+(.*\S)\s*$", line)
        if m:
            headings.append((i, len(m.group(1)), m.group(2).strip()))
    if not headings:
        return {}

    levels = sorted({lvl for _, lvl, _ in headings})
    top_level = levels[0]
    top_count = sum(1 for _, lvl, _ in headings if lvl == top_level)
    # A single top-level heading is the document title -> sections live one level down.
    if top_count == 1 and len(levels) > 1:
        section_level = levels[1]
    else:
        section_level = top_level

    section_heads = [(i, text) for (i, lvl, text) in headings if lvl == section_level]
    out: dict[str, tuple[str, str]] = {}
    for k, (idx, text) in enumerate(section_heads):
        end = section_heads[k + 1][0] if k + 1 < len(section_heads) else len(lines)
        body = "\n".join(lines[idx:end]).strip()
        norm = _normalize_heading(text)
        if norm and norm not in out:
            out[norm] = (text, body)
    return out


def _match_sample_section(title: str, sample_sections: dict[str, tuple[str, str]]) -> str | None:
    """Match a requested section title to one sample section block, or None.

    Tries an exact heading match first, then a normalized match.
    """
    if not sample_sections:
        return None
    for _norm, (orig, body) in sample_sections.items():
        if orig.strip() == title.strip():
            return body
    match = sample_sections.get(_normalize_heading(title))
    return match[1] if match else None


# ---------------------------------------------------------------------------
# Main module
# ---------------------------------------------------------------------------


class MultiSourceReportGenerator:
    """Generate a user-defined Markdown report from many mixed source files.

    Example
    -------
    >>> gen = MultiSourceReportGenerator(use_azure=True)
    >>> result = gen.run(
    ...     input_paths=["materials/"],
    ...     report_title="Project Assessment",
    ...     sections=[{"title": "Findings", "instructions": "Summarize key findings."}],
    ...     output_dir="output/report",
    ... )
    >>> print(result.markdown_path)
    """

    def __init__(self, *, api_config: dict | None = None, use_azure: bool = True) -> None:
        if api_config is None:
            from gaik.software_components.config import get_openai_config

            api_config = get_openai_config(use_azure=use_azure)
        self.api_config = api_config

    # -- public API -------------------------------------------------------

    def run(
        self,
        *,
        input_paths: list[str | Path],
        sections: list[ReportSectionSpec | dict[str, Any]],
        report_title: str = "Generated Report",
        report_description: str | None = None,
        report_language: str | None = None,
        sample_report_path: str | Path | None = None,
        output_dir: str | Path | None = None,
        include_evidence_index: bool = True,
        include_source_references: bool = True,
        max_evidence_chars: int | None = None,
        section_context_mode: str = "all_evidence",
        parser_choice: str = "auto",
        parser_options: dict | None = None,
        transcriber_options: dict | None = None,
        image_options: dict | None = None,
        writer_options: dict | None = None,
        agentic: bool = False,
        review_options: dict | None = None,
        polish: bool = False,
        strict_review: bool = False,
        curate_evidence: bool = False,
        verbose: bool = False,
        progress_callback: Callable[[str], None] | None = None,
        output_docx: bool = False,
    ) -> ReportGenerationResult:
        """Build evidence from ``input_paths`` and write each requested section.

        By default the whole report is written in a single LLM call. Set
        ``agentic=True`` to use the opt-in agentic workflow (independent
        per-section drafting + mandatory diff-editor review); it requires the
        ``multi-source-report-generator-agentic`` extra (``langgraph``).
        """
        specs = _normalize_sections(sections)

        parser_options = parser_options or {}
        transcriber_options = transcriber_options or {}
        image_options = image_options or {}
        writer_options = writer_options or {}

        # 1. Build evidence (shared by both paths).
        evidence_items, evidence_pack, sample_markdown = self._build_evidence(
            input_paths=input_paths,
            parser_choice=parser_choice,
            parser_options=parser_options,
            transcriber_options=transcriber_options,
            image_options=image_options,
            sample_report_path=sample_report_path,
            max_evidence_chars=max_evidence_chars,
            progress_callback=progress_callback,
        )

        # Suppress inline citations when there is only one source — nothing
        # to distinguish between sources, so citations add noise.
        if len(evidence_items) <= 1:
            include_source_references = False

        source_filenames = [item.metadata.get("filename", "") for item in evidence_items]

        # 2. Write the report — either the agentic path or the single LLM call.
        if agentic:
            markdown, generated, usage_total = self._run_agentic(
                specs=specs,
                evidence_items=evidence_items,
                evidence_pack=evidence_pack,
                sample_markdown=sample_markdown,
                output_dir=output_dir,
                report_title=report_title,
                report_description=report_description,
                report_language=report_language,
                include_source_references=include_source_references,
                source_filenames=source_filenames,
                writer_options=writer_options,
                review_options=review_options,
                polish=polish,
                strict_review=strict_review,
                curate_evidence=curate_evidence,
                verbose=verbose,
                progress_callback=progress_callback,
            )
        else:
            # Write the whole report in a single LLM call (so a sample report's
            # format applies to the report as a whole, not per section).
            client = self._build_llm_client(writer_options)
            chat_kwargs = {
                k: v
                for k, v in writer_options.items()
                if k not in ("model", "provider") and v is not None
            }

            response = self._write_report(
                client,
                specs=specs,
                evidence_pack=evidence_pack,
                report_title=report_title,
                report_description=report_description,
                report_language=report_language,
                include_source_references=include_source_references,
                source_filenames=source_filenames,
                sample_markdown=sample_markdown,
                chat_kwargs=chat_kwargs,
            )
            markdown = _normalize_report_markdown(response.text, report_title)
            usage_total = (
                {k: v for k, v in response.usage.items() if isinstance(v, int)}
                if getattr(response, "usage", None)
                else {}
            )
            generated = _split_into_sections(markdown, specs)

        # 3. Write outputs
        markdown_path: Path | None = None
        docx_path: Path | None = None
        if output_dir is not None:
            markdown_path = self._write_outputs(
                output_dir,
                report_title=report_title,
                markdown=markdown,
                sections=generated,
                evidence_items=evidence_items,
                evidence_pack=evidence_pack,
                usage=usage_total,
                include_evidence_index=include_evidence_index,
            )
            if output_docx and markdown_path is not None:
                docx_path = self._write_docx(markdown_path)

        return ReportGenerationResult(
            title=report_title,
            evidence_items=evidence_items,
            sections=generated,
            markdown=markdown,
            markdown_path=markdown_path,
            usage=usage_total,
            docx_path=docx_path,
        )

    # -- shared evidence building -----------------------------------------

    def _build_evidence(
        self,
        *,
        input_paths: list[str | Path],
        parser_choice: str,
        parser_options: dict,
        transcriber_options: dict,
        image_options: dict,
        sample_report_path: str | Path | None,
        max_evidence_chars: int | None,
        progress_callback: Callable[[str], None] | None = None,
    ) -> tuple[list[EvidenceItem], str, str | None]:
        """Normalize all inputs to an evidence pack (+ optional sample markdown).

        Shared by the single-call and agentic paths. When ``progress_callback``
        is provided, normalization progress is routed through it instead of
        ``print`` so callers (e.g. the demo SSE backend) receive the messages.
        """
        _emit = progress_callback or print

        files = _collect_input_files(input_paths)
        if not files:
            raise ValueError(
                "No supported source files found in `input_paths`. "
                f"Supported extensions: {sorted(SUPPORTED_EXTENSIONS)}"
            )

        total = len(files)
        evidence_items: list[EvidenceItem] = []
        for index, path in enumerate(files, start=1):
            source_type = _EXT_TO_TYPE[path.suffix.lower()]
            _emit(f"Normalizing [{index}/{total}]: {path.name} ({source_type})")
            evidence_items.append(
                self._build_evidence_item(
                    path,
                    index=index,
                    parser_choice=parser_choice,
                    parser_options=parser_options,
                    transcriber_options=transcriber_options,
                    image_options=image_options,
                )
            )

        evidence_pack = self._assemble_evidence_pack(
            evidence_items, max_evidence_chars=max_evidence_chars
        )
        _emit(f"Evidence pack assembled: {total} source(s), {len(evidence_pack):,} chars")

        sample_markdown: str | None = None
        if sample_report_path is not None:
            sample_markdown = self._extract_sample_report(
                sample_report_path,
                parser_choice=parser_choice,
                parser_options=parser_options,
            )
            if max_evidence_chars is not None and len(sample_markdown) > max_evidence_chars:
                sample_markdown = (
                    sample_markdown[:max_evidence_chars] + "\n\n[... sample report truncated ...]"
                )

        return evidence_items, evidence_pack, sample_markdown

    # -- agentic path -----------------------------------------------------

    def _run_agentic(
        self,
        *,
        specs: list[ReportSectionSpec],
        evidence_items: list[EvidenceItem],
        evidence_pack: str,
        sample_markdown: str | None,
        output_dir: str | Path | None,
        report_title: str,
        report_description: str | None,
        report_language: str | None,
        include_source_references: bool,
        source_filenames: list[str],
        writer_options: dict,
        review_options: dict | None,
        polish: bool,
        strict_review: bool,
        curate_evidence: bool,
        verbose: bool,
        progress_callback: Callable[[str], None] | None,
    ) -> tuple[str, list[GeneratedSection], dict]:
        from .agentic import run_agentic_report
        from .progress import ProgressReporter

        reporter = ProgressReporter(verbose=verbose, callback=progress_callback)

        writer_client = self._build_llm_client(writer_options)
        writer_kwargs = {
            k: v
            for k, v in writer_options.items()
            if k not in ("model", "provider") and v is not None
        }

        # Reviewer client: a separate model via review_options, else reuse the writer.
        if review_options:
            reviewer_client = self._build_llm_client(review_options)
            reviewer_kwargs = {
                k: v
                for k, v in review_options.items()
                if k not in ("model", "provider") and v is not None
            }
        else:
            reviewer_client = writer_client
            reviewer_kwargs = writer_kwargs

        # Split the sample by heading and resolve a matched block per section
        # here (so the agentic package never imports pipeline internals).
        sample_sections = _split_sample_sections(sample_markdown) if sample_markdown else {}
        matched_samples = {
            spec.title: _match_sample_section(spec.title, sample_sections) for spec in specs
        }

        markdown, generated, usage_total = run_agentic_report(
            specs=specs,
            evidence_items=evidence_items,
            evidence_pack=evidence_pack,
            matched_samples=matched_samples,
            report_description=report_description,
            sample_report_provided=sample_markdown is not None,
            output_dir=output_dir,
            report_title=report_title,
            report_language=report_language,
            include_source_references=include_source_references,
            source_filenames=source_filenames,
            writer_client=writer_client,
            writer_kwargs=writer_kwargs,
            reviewer_client=reviewer_client,
            reviewer_kwargs=reviewer_kwargs,
            curate_evidence=curate_evidence,
            polish=polish,
            strict_review=strict_review,
            reporter=reporter,
        )
        return _normalize_report_markdown(markdown, report_title), generated, usage_total

    # -- evidence building ------------------------------------------------

    def _build_evidence_item(
        self,
        path: Path,
        *,
        index: int,
        parser_choice: str,
        parser_options: dict,
        transcriber_options: dict,
        image_options: dict,
    ) -> EvidenceItem:
        source_type = _EXT_TO_TYPE[path.suffix.lower()]
        content = self._extract_content(
            path,
            source_type=source_type,
            parser_choice=parser_choice,
            parser_options=parser_options,
            transcriber_options=transcriber_options,
            image_options=image_options,
        )
        metadata = {
            "index": index,
            "filename": path.name,
            "extension": path.suffix.lower(),
            "detected_type": source_type,
            "size_bytes": path.stat().st_size if path.exists() else None,
        }
        return EvidenceItem(
            source_path=str(path),
            source_type=source_type,
            content_markdown=content,
            metadata=metadata,
        )

    def _extract_content(
        self,
        path: Path,
        *,
        source_type: str,
        parser_choice: str,
        parser_options: dict,
        transcriber_options: dict,
        image_options: dict,
    ) -> str:
        if source_type in ("text", "markdown"):
            return path.read_text(encoding="utf-8", errors="replace")
        if source_type == "csv":
            return _csv_to_markdown(path)
        if source_type == "xlsx":
            return _xlsx_to_markdown(path)
        if source_type == "pdf":
            return self._parse_pdf(path, parser_choice=parser_choice, parser_options=parser_options)
        if source_type == "docx":
            return self._parse_docx(path)
        if source_type in ("audio", "video"):
            return self._transcribe(path, transcriber_options=transcriber_options)
        if source_type == "image":
            return self._parse_image(path, image_options=image_options)
        # Should never happen — _collect_input_files filters unsupported types.
        raise ValueError(f"Unsupported source type '{source_type}' for {path}")

    def _extract_sample_report(
        self, sample_report_path: str | Path, *, parser_choice: str, parser_options: dict
    ) -> str:
        """Parse an optional sample/template report to markdown.

        Supports text, Markdown, PDF, and DOCX. The result is used as a strict
        format/style template by the writer — never as content.
        """
        path = Path(sample_report_path)
        if not path.exists():
            raise FileNotFoundError(f"sample_report_path does not exist: {path}")
        ext = path.suffix.lower()
        source_type = _EXT_TO_TYPE.get(ext)
        if source_type not in ("text", "markdown", "pdf", "docx"):
            raise ValueError(
                f"Unsupported sample report type '{ext}'. "
                "Use one of: .txt, .md, .markdown, .pdf, .docx"
            )
        return self._extract_content(
            path,
            source_type=source_type,
            parser_choice=parser_choice,
            parser_options=parser_options,
            transcriber_options={},
            image_options={},
        )

    def _parse_pdf(self, path: Path, *, parser_choice: str, parser_options: dict) -> str:
        choice = (parser_choice or "auto").lower()
        if choice == "auto":
            choice = "pymupdf"  # local, no API cost; deterministic default for V1

        if choice == "pymupdf":
            from gaik.software_components.parsers import PyMuPDFParser

            return PyMuPDFParser().parse_pdf(str(path), use_markdown=True)
        if choice in ("vision", "vision_parser"):
            from gaik.software_components.parsers import VisionParser, get_openai_config

            cfg = parser_options.get("openai_config") or get_openai_config(
                use_azure=self.api_config.get("use_azure", True)
            )
            pages = VisionParser(cfg, **parser_options.get("ctor", {})).convert_pdf(str(path))
            return "\n\n".join(pages)
        if choice == "multimodal":
            from gaik.software_components.parsers import MultimodalParser

            result = MultimodalParser(**parser_options.get("ctor", {})).parse(str(path))
            return result.clean_markdown or result.raw_markdown
        if choice == "docling":
            from gaik.software_components.parsers import DoclingParser

            result = DoclingParser(**parser_options.get("ctor", {})).parse_document(str(path))
            return _coerce_markdown(result)
        raise ValueError(
            f"Unknown parser_choice '{parser_choice}'. "
            "Use one of: auto, pymupdf, vision, multimodal, docling."
        )

    def _parse_docx(self, path: Path) -> str:
        from gaik.software_components.parsers import DocxParser

        return DocxParser().parse_docx(str(path), use_markdown=True)

    def _transcribe(self, path: Path, *, transcriber_options: dict) -> str:
        from gaik.software_components.transcriber import Transcriber

        ctor = dict(transcriber_options.get("ctor", {}))
        transcriber = Transcriber(api_config=self.api_config, **ctor)
        result = transcriber.transcribe(str(path), **transcriber_options.get("call", {}))
        return result.enhanced_transcript or result.raw_transcript

    def _parse_image(self, path: Path, *, image_options: dict) -> str:
        mode = image_options.get("mode", "parse")  # "parse" | "structured"
        if mode == "structured":
            from gaik.software_components.vision_extractor import VisionExtractor

            ctor = dict(image_options.get("ctor", {}))
            ctor.setdefault("api_config", self.api_config)
            extractor = VisionExtractor(**ctor)
            user_requirements = image_options.get(
                "user_requirements",
                "Extract all visible text, tables, figures, and structured content from the image.",
            )
            result = extractor.extract(file_paths=[str(path)], user_requirements=user_requirements)
            return _dict_to_markdown(result.data)

        # default: general parsing to markdown via VisionParser.convert_image()
        from gaik.software_components.parsers import VisionParser, get_openai_config

        cfg = image_options.get("openai_config") or get_openai_config(
            use_azure=self.api_config.get("use_azure", True)
        )
        return VisionParser(cfg, **image_options.get("ctor", {})).convert_image(str(path))

    def _assemble_evidence_pack(
        self, evidence_items: list[EvidenceItem], *, max_evidence_chars: int | None
    ) -> str:
        blocks: list[str] = []
        for item in evidence_items:
            meta = item.metadata
            header = (
                f"## Source {meta.get('index')}: {meta.get('filename')}\n"
                f"Type: {item.source_type}\n"
                f"Path: {item.source_path}\n"
            )
            blocks.append(f"{header}\n{item.content_markdown}")
        pack = "\n\n---\n\n".join(blocks)
        if max_evidence_chars is not None and len(pack) > max_evidence_chars:
            logger.warning(
                "Evidence pack (%d chars) exceeds max_evidence_chars (%d); truncating.",
                len(pack),
                max_evidence_chars,
            )
            pack = pack[:max_evidence_chars] + "\n\n[... evidence truncated ...]"
        return pack

    # -- writing ----------------------------------------------------------

    def _build_llm_client(self, writer_options: dict):
        if create_llm_client is None:
            raise ImportError(
                "The LLM client could not be imported. Install the base GAIK "
                "dependencies (openai) to run the report writer."
            )
        cfg = dict(self.api_config)
        if writer_options.get("model"):
            cfg["model"] = writer_options["model"]
        if writer_options.get("provider"):
            cfg["provider"] = writer_options["provider"]
        return create_llm_client(cfg)

    def _write_report(
        self,
        client,
        *,
        specs: list[ReportSectionSpec],
        evidence_pack: str,
        report_title: str,
        report_description: str | None,
        report_language: str | None,
        include_source_references: bool,
        source_filenames: list[str],
        sample_markdown: str | None,
        chat_kwargs: dict,
    ):
        """Write the entire report in one LLM call and return the raw response."""
        parts = [f"Report title: {report_title}"]
        if report_description:
            parts.append(f"Report context and purpose: {report_description}")
        if report_language:
            parts.append(f"Write the report in: {report_language}")
        if include_source_references:
            parts.append(
                "Where useful, cite the source that supports a claim using its exact filename "
                "in parentheses, e.g. (notes.txt) or (meeting_recording.mp3). "
                f"Available sources: {', '.join(source_filenames)}."
            )
        else:
            parts.append("Do not add inline source citations or filename references in the text.")

        # Present the sample BEFORE the section instructions so the model internalises
        # the expected tone/structure/format (layout, list style, brevity) before
        # reading what content to put in each section.
        if sample_markdown:
            parts.append(
                "FORMAT REFERENCE — governs both structure AND length for every section. "
                "This example is most likely about a DIFFERENT subject. "
                "Mirror it exactly:\n"
                "  • Structure: how each section is laid out internally, list style "
                "(prose/bullets/numbered), bold lead-ins, citation style.\n"
                "  • Length: each section you write should be approximately the same "
                "total length and density as the corresponding section in this reference "
                "— same number of paragraphs or bullet points, similar per-item length. "
                "Do not add extra content to be thorough; do not pad.\n"
                "  • Content: take NONE of its facts, names, numbers, or wording. Every "
                "fact must come from the EVIDENCE below. Section titles come from the "
                "required list.\n\n"
                f"{sample_markdown}"
            )

        # Section definitions: use a numbered, titled format so the model never
        # echoes a "SECTION:" prefix into the output headings.
        section_lines = ["Required sections (write each as `## <title>` in this exact order):"]
        for i, spec in enumerate(specs, start=1):
            section_lines.append(f"\n{i}. Heading: {spec.title}")
            section_lines.append(f"   Content to cover: {spec.instructions}")
        parts.append("\n".join(section_lines))

        parts.append(
            "Evidence — this is the ONLY source of facts and content for the report:\n"
            f"{evidence_pack}"
        )

        closing = [
            f"Now write the complete report. Start with `# {report_title}`, then write "
            "every required section in order using `## <exact heading title>`."
        ]
        if sample_markdown:
            closing.append(
                "Follow the FORMAT REFERENCE's layout, style, AND length — each section "
                "should be approximately as long as the corresponding reference section. "
                "Ignore its subject matter entirely; all facts come from the evidence."
            )
        else:
            closing.append("Write in a clean, professional report format.")
        closing.append("Use only the evidence for content.")
        parts.append(" ".join(closing))
        user_prompt = "\n\n".join(parts)

        return client.chat(
            [
                {"role": "system", "content": _SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            **chat_kwargs,
        )

    # -- outputs ----------------------------------------------------------

    def _write_outputs(
        self,
        output_dir: str | Path,
        *,
        report_title: str,
        markdown: str,
        sections: list[GeneratedSection],
        evidence_items: list[EvidenceItem],
        evidence_pack: str,
        usage: dict[str, int],
        include_evidence_index: bool,
    ) -> Path:
        out = Path(output_dir)
        out.mkdir(parents=True, exist_ok=True)
        (out / "sections").mkdir(exist_ok=True)
        (out / "evidence").mkdir(exist_ok=True)

        report_path = out / "report.md"
        report_path.write_text(markdown, encoding="utf-8")

        for i, s in enumerate(sections, start=1):
            name = f"{i:02d}_{_slug(s.title)}.md"
            (out / "sections" / name).write_text(
                f"## {s.title}\n\n{s.content_markdown.strip()}\n", encoding="utf-8"
            )

        (out / "evidence" / "normalized_sources.md").write_text(evidence_pack, encoding="utf-8")

        if include_evidence_index:
            index = [
                {
                    "source_path": e.source_path,
                    "source_type": e.source_type,
                    "metadata": e.metadata,
                    "content_chars": len(e.content_markdown),
                }
                for e in evidence_items
            ]
            (out / "evidence_index.json").write_text(
                json.dumps(index, indent=2, ensure_ascii=False), encoding="utf-8"
            )

        if usage:
            (out / "usage.json").write_text(
                json.dumps(usage, indent=2, ensure_ascii=False), encoding="utf-8"
            )

        return report_path

    def _write_docx(self, md_path: Path) -> Path:
        """Convert a Markdown file to DOCX using Pandoc via pypandoc.

        Requires the ``multi-source-report-generator-docx`` extra and the
        Pandoc system binary (https://pandoc.org/installing.html).
        """
        try:
            import pypandoc
        except ImportError as exc:
            raise ImportError(
                "DOCX export requires 'pypandoc' and the Pandoc system binary. "
                "Install the Python wrapper with:\n"
                '    pip install "gaik[multi-source-report-generator-docx]"\n'
                "Then install Pandoc from https://pandoc.org/installing.html "
                "(or: winget install JohnMacFarlane.Pandoc / brew install pandoc / "
                "apt install pandoc)."
            ) from exc
        docx_path = md_path.with_suffix(".docx")
        pypandoc.convert_file(str(md_path), "docx", outputfile=str(docx_path))
        return docx_path


# ---------------------------------------------------------------------------
# Config save / load
# ---------------------------------------------------------------------------


def save_report_config(
    path: str | Path,
    *,
    input_paths: list[str | Path],
    sections: list[ReportSectionSpec | dict[str, Any]],
    report_title: str = "Generated Report",
    report_description: str | None = None,
    report_language: str | None = None,
    sample_report_path: str | Path | None = None,
    output_dir: str | Path | None = None,
    output_docx: bool = False,
    include_evidence_index: bool = True,
    include_source_references: bool = True,
    max_evidence_chars: int | None = None,
    parser_choice: str = "auto",
    parser_options: dict | None = None,
    transcriber_options: dict | None = None,
    image_options: dict | None = None,
    writer_options: dict | None = None,
    agentic: bool = False,
    review_options: dict | None = None,
    polish: bool = False,
    strict_review: bool = False,
    curate_evidence: bool = False,
) -> Path:
    """Persist all ``run()`` parameters to a JSON config file for later reuse.

    Paths (``input_paths``, ``output_dir``, ``sample_report_path``) are stored
    relative to the config file's directory so the config is portable. All keys
    inside option dicts (``transcriber_options``, ``parser_options``, etc.) are
    stored as-is — any option supported by ``run()`` is preserved.

    Not persisted: ``verbose``, ``progress_callback`` (runtime display), and
    ``section_context_mode`` (reserved/unused). Pass them directly to ``run()``
    as needed.
    """
    config_path = Path(path)
    base = config_path.parent

    def _to_rel(p: str | Path | None) -> str | None:
        if p is None:
            return None
        try:
            return str(Path(p).relative_to(base.resolve()))
        except ValueError:
            return str(p)  # already absolute or on a different drive

    # Normalise sections: dataclasses → plain dicts, preserve all fields.
    section_dicts: list[dict[str, Any]] = []
    for s in sections:
        if isinstance(s, ReportSectionSpec):
            section_dicts.append(
                {
                    "id": s.id,
                    "title": s.title,
                    "instructions": s.instructions,
                    "required": s.required,
                    "depends_on": list(s.depends_on),
                }
            )
        else:
            section_dicts.append(dict(s))

    config: dict[str, Any] = {
        "version": "1",
        "report_title": report_title,
        "report_description": report_description,
        "report_language": report_language,
        "sections": section_dicts,
        "input_paths": [_to_rel(p) for p in input_paths],
        "sample_report_path": _to_rel(sample_report_path),
        "output_dir": _to_rel(output_dir),
        "output_docx": output_docx,
        "include_evidence_index": include_evidence_index,
        "include_source_references": include_source_references,
        "max_evidence_chars": max_evidence_chars,
        "parser_choice": parser_choice,
        "parser_options": parser_options or {},
        "transcriber_options": transcriber_options or {},
        "image_options": image_options or {},
        "writer_options": writer_options or {},
        "agentic": agentic,
        "review_options": review_options,
        "polish": polish,
        "strict_review": strict_review,
        "curate_evidence": curate_evidence,
    }

    config_path.parent.mkdir(parents=True, exist_ok=True)
    config_path.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    return config_path


def load_report_config(path: str | Path) -> dict[str, Any]:
    """Load a config file saved by ``save_report_config`` and return a dict
    that can be unpacked directly into ``MultiSourceReportGenerator.run()``.

    Relative paths stored in the config are resolved relative to the config
    file's directory. ``None`` is returned for any key that was not present.
    """
    config_path = Path(path)
    base = config_path.parent

    raw: dict[str, Any] = json.loads(config_path.read_text(encoding="utf-8"))

    def _to_abs(p: str | None) -> Path | None:
        if p is None:
            return None
        resolved = Path(p)
        return resolved if resolved.is_absolute() else (base / resolved).resolve()

    return {
        "input_paths": [_to_abs(p) for p in raw.get("input_paths") or []],
        "sections": raw.get("sections", []),
        "report_title": raw.get("report_title", "Generated Report"),
        "report_description": raw.get("report_description"),
        "report_language": raw.get("report_language"),
        "sample_report_path": _to_abs(raw.get("sample_report_path")),
        "output_dir": _to_abs(raw.get("output_dir")),
        "output_docx": raw.get("output_docx", False),
        "include_evidence_index": raw.get("include_evidence_index", True),
        "include_source_references": raw.get("include_source_references", True),
        "max_evidence_chars": raw.get("max_evidence_chars"),
        "parser_choice": raw.get("parser_choice", "auto"),
        "parser_options": raw.get("parser_options"),
        "transcriber_options": raw.get("transcriber_options"),
        "image_options": raw.get("image_options"),
        "writer_options": raw.get("writer_options"),
        "agentic": raw.get("agentic", False),
        "review_options": raw.get("review_options"),
        "polish": raw.get("polish", False),
        "strict_review": raw.get("strict_review", False),
        "curate_evidence": raw.get("curate_evidence", False),
    }


# ---------------------------------------------------------------------------
# Markdown coercion helpers
# ---------------------------------------------------------------------------


def _coerce_markdown(result: Any) -> str:
    """Best-effort extraction of markdown text from a parser result object."""
    for attr in ("clean_markdown", "parsed_markdown", "markdown", "text"):
        val = getattr(result, attr, None)
        if isinstance(val, str) and val:
            return val
    if isinstance(result, dict):
        for key in ("clean_markdown", "parsed_markdown", "markdown", "text"):
            if isinstance(result.get(key), str):
                return result[key]
    if isinstance(result, str):
        return result
    return str(result)


def _dict_to_markdown(data: Any) -> str:
    """Serialize a VisionExtractor result dict into readable markdown evidence."""
    if isinstance(data, dict) and data:
        lines = []
        for key, value in data.items():
            label = str(key).replace("_", " ").strip().title()
            if isinstance(value, (dict, list)):
                rendered = json.dumps(value, indent=2, ensure_ascii=False)
                lines.append(f"**{label}:**\n```json\n{rendered}\n```")
            else:
                lines.append(f"**{label}:** {value}")
        return "\n\n".join(lines)
    # Fallback: fenced JSON block
    return f"```json\n{json.dumps(data, indent=2, ensure_ascii=False)}\n```"
