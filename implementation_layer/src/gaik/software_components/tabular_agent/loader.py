"""Turn user files into DuckDB tables, then lock the connection down.

CSV, TSV, Parquet and JSON go straight through DuckDB's own readers, which
handle delimiter sniffing, encodings and type inference far better than
anything worth hand-rolling.

Excel is the hard case, because real spreadsheets are laid out for humans: a
title in row 1, blank spacer rows, headers spread over merged cells, subtotal
rows in the middle, a "Notes:" block at the bottom. Two tiers handle that:

1. **Heuristics** (always, no LLM call): drop blank edges, find the header row
   by looking for the first mostly-populated, mostly-textual row, and name the
   columns. Clean exports never get past this tier.
2. **LLM layout inference** (only when tier 1 flags the sheet as messy): the
   raw cell grid is shown to the model, which returns a :class:`SheetLayout`
   saying where the table actually starts and ends.

Normalized sheets are written to a temporary CSV and read back through DuckDB,
so Excel and CSV inputs end up with identical type inference.
"""

from __future__ import annotations

import csv
import logging
import re
import tempfile
from collections.abc import Callable, Mapping, Sequence
from pathlib import Path
from typing import Any

from .models import SheetLayout

logger = logging.getLogger(__name__)

CSV_EXTENSIONS = {".csv", ".tsv", ".txt"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
PARQUET_EXTENSIONS = {".parquet"}
JSON_EXTENSIONS = {".json", ".ndjson"}
SUPPORTED_EXTENSIONS = CSV_EXTENSIONS | EXCEL_EXTENSIONS | PARQUET_EXTENSIONS | JSON_EXTENSIONS

# How much of a sheet is shown to the LLM when inferring layout.
PREVIEW_ROWS = 30
PREVIEW_COLS = 15
# Rows scanned by the heuristic when hunting for the header.
_HEADER_SEARCH_ROWS = 25

_IDENTIFIER_SAFE = re.compile(r"[^0-9a-zA-Z_]+")

LayoutInferrer = Callable[[str, str], SheetLayout]
"""Callback: (sheet description, rendered cell grid) -> SheetLayout."""


class TableLoadError(ValueError):
    """Raised when a source file cannot be turned into a table."""


def sanitize_identifier(name: str, *, fallback: str = "table") -> str:
    """Turn an arbitrary filename or sheet name into a safe SQL identifier."""
    cleaned = _IDENTIFIER_SAFE.sub("_", (name or "").strip()).strip("_").lower()
    if not cleaned:
        cleaned = fallback
    if cleaned[0].isdigit():
        cleaned = f"t_{cleaned}"
    return cleaned[:60]


def _unique(name: str, taken: set[str]) -> str:
    """Append a numeric suffix until ``name`` is unused."""
    if name not in taken:
        return name
    for suffix in range(2, 1000):
        candidate = f"{name}_{suffix}"
        if candidate not in taken:
            return candidate
    raise TableLoadError(f"Could not find a unique table name for '{name}'.")


def normalize_sources(
    source: str | Path | Sequence[str | Path] | Mapping[str, str | Path],
) -> list[tuple[str | None, Path]]:
    """Normalize the constructor's ``source`` into ``(explicit_name, path)`` pairs.

    Args:
        source: A path, a sequence of paths, or a ``{table_name: path}`` mapping.

    Returns:
        Pairs whose first element is the caller's chosen table name, or ``None``
        to derive one from the filename.

    Raises:
        TableLoadError: If ``source`` is empty, or a path does not exist.
    """
    pairs: list[tuple[str | None, Path]] = []
    if isinstance(source, Mapping):
        pairs = [(str(name), Path(path)) for name, path in source.items()]
    elif isinstance(source, (str, Path)):
        pairs = [(None, Path(source))]
    elif isinstance(source, Sequence):
        pairs = [(None, Path(p)) for p in source]
    else:
        raise TableLoadError(f"Unsupported source type: {type(source).__name__}")

    if not pairs:
        raise TableLoadError("No source files given.")

    for _, path in pairs:
        if not path.exists():
            raise TableLoadError(f"File not found: {path}")
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            raise TableLoadError(
                f"Unsupported file type '{path.suffix}' for {path.name}. "
                f"Supported: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
            )
    return pairs


def load_sources(
    con: Any,
    sources: list[tuple[str | None, Path]],
    *,
    layout_inference: str = "auto",
    infer_layout: LayoutInferrer | None = None,
) -> dict[str, str]:
    """Load every source file into ``con`` as a table.

    Args:
        con: An open DuckDB connection, not yet locked down.
        sources: ``(explicit_name, path)`` pairs from :func:`normalize_sources`.
        layout_inference: ``"auto"``, ``"always"`` or ``"never"`` -- when the LLM
            may be asked to work out a messy sheet's layout.
        infer_layout: Callback invoked for that inference. When ``None``, the
            deterministic heuristics are used no matter what.

    Returns:
        A ``{table_name: human-readable source description}`` mapping.

    Raises:
        TableLoadError: If a file cannot be read or produces no usable table.
    """
    created: dict[str, str] = {}
    taken: set[str] = set()

    for explicit_name, path in sources:
        suffix = path.suffix.lower()
        if suffix in EXCEL_EXTENSIONS:
            sheets = _load_excel(
                con,
                path,
                explicit_name=explicit_name,
                taken=taken,
                layout_inference=layout_inference,
                infer_layout=infer_layout,
            )
            created.update(sheets)
        else:
            base = sanitize_identifier(explicit_name or path.stem)
            name = _unique(base, taken)
            taken.add(name)
            _load_flat_file(con, path, name)
            created[name] = path.name

    if not created:
        raise TableLoadError("No tables could be loaded from the given files.")
    return created


def _load_flat_file(con: Any, path: Path, table_name: str) -> None:
    """Create a table from a CSV/TSV/Parquet/JSON file using DuckDB's readers."""
    suffix = path.suffix.lower()
    posix = path.resolve().as_posix()
    if suffix in PARQUET_EXTENSIONS:
        reader = f"read_parquet('{_escape(posix)}')"
    elif suffix in JSON_EXTENSIONS:
        reader = f"read_json_auto('{_escape(posix)}')"
    else:
        reader = f"read_csv('{_escape(posix)}', auto_detect=true, sample_size=-1)"
    try:
        con.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM {reader}')
    except Exception as exc:
        raise TableLoadError(f"Could not read {path.name}: {exc}") from exc
    if suffix not in PARQUET_EXTENSIONS:
        repair_european_decimals(con, table_name)


# "1234,56" or "1.234.567,89" -- a decimal comma, optionally with dot as the
# thousands separator. Anchored, so free text containing a comma never matches.
_EURO_DECIMAL = r"^-?[0-9]{1,3}(\.[0-9]{3})*,[0-9]+$|^-?[0-9]+,[0-9]+$"


def repair_european_decimals(con: Any, table_name: str) -> list[str]:
    """Retype text columns that hold comma-decimal numbers as DOUBLE.

    DuckDB's type inference reads ``1234,56`` as text, so a Finnish or Swedish
    export silently loses every numeric column and no amount of prompting makes
    ``SUM()`` work on it. A column is converted only when *every* non-null value
    is a comma-decimal number, which leaves genuine free text untouched.

    Returns:
        The names of the columns that were converted.
    """
    try:
        described = con.execute(f'DESCRIBE "{table_name}"').fetchall()
    except Exception:  # pragma: no cover - table always exists at this point
        return []

    converted: list[str] = []
    for row in described:
        column, data_type = row[0], str(row[1]).upper()
        if not data_type.startswith("VARCHAR"):
            continue
        quoted = f'"{column}"'
        try:
            total, matching = con.execute(
                f"SELECT count({quoted}), "
                f"count(*) FILTER (WHERE regexp_matches({quoted}, ?)) "
                f'FROM "{table_name}"',
                [_EURO_DECIMAL],
            ).fetchone()
        except Exception as exc:  # pragma: no cover - defensive
            logger.debug("Decimal probe failed for %s.%s: %s", table_name, column, exc)
            continue
        if not total or matching != total:
            continue
        try:
            con.execute(
                f'ALTER TABLE "{table_name}" ALTER {quoted} TYPE DOUBLE '
                f"USING CAST(replace(replace({quoted}, '.', ''), ',', '.') AS DOUBLE)"
            )
            converted.append(column)
        except Exception as exc:
            logger.debug("Could not retype %s.%s as DOUBLE: %s", table_name, column, exc)
    if converted:
        logger.info(
            "Converted comma-decimal column(s) in '%s': %s", table_name, ", ".join(converted)
        )
    return converted


def _escape(text: str) -> str:
    """Escape single quotes for a DuckDB string literal."""
    return text.replace("'", "''")


def _load_excel(
    con: Any,
    path: Path,
    *,
    explicit_name: str | None,
    taken: set[str],
    layout_inference: str,
    infer_layout: LayoutInferrer | None,
) -> dict[str, str]:
    """Create one table per non-empty worksheet in an Excel workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading Excel files requires 'openpyxl'. "
            "Install extras with 'pip install gaik[tabular-agent]'"
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise TableLoadError(f"Could not open {path.name}: {exc}") from exc

    created: dict[str, str] = {}
    try:
        sheet_names = list(workbook.sheetnames)
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            grid = _read_grid(worksheet)
            grid = _trim_grid(grid)
            if len(grid) < 2:
                logger.info("Skipping empty sheet '%s' in %s", sheet_name, path.name)
                continue

            layout, messy = _resolve_layout(
                grid,
                sheet_label=f"{path.name} / {sheet_name}",
                layout_inference=layout_inference,
                infer_layout=infer_layout,
            )
            header, rows = _apply_layout(grid, layout, drop_outliers=messy)
            if not header or not rows:
                logger.info("Sheet '%s' in %s has no data rows", sheet_name, path.name)
                continue

            # One sheet keeps the caller's name; extra sheets get suffixed.
            if explicit_name and len(sheet_names) == 1:
                base = sanitize_identifier(explicit_name)
            elif explicit_name:
                base = sanitize_identifier(f"{explicit_name}_{sheet_name}")
            else:
                base = sanitize_identifier(
                    sheet_name if len(sheet_names) > 1 else path.stem,
                    fallback=sanitize_identifier(path.stem),
                )
            name = _unique(base, taken)
            taken.add(name)

            _create_from_rows(con, name, header, rows)
            created[name] = f"{path.name} (sheet '{sheet_name}')"
    finally:
        workbook.close()

    if not created:
        raise TableLoadError(f"No usable sheets found in {path.name}.")
    return created


def _read_grid(worksheet: Any) -> list[list[Any]]:
    """Read a worksheet into a rectangular list-of-lists of raw cell values."""
    grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
    width = max((len(r) for r in grid), default=0)
    return [list(r) + [None] * (width - len(r)) for r in grid]


def _is_blank(value: Any) -> bool:
    """True for cells that hold nothing a human would call a value."""
    return value is None or (isinstance(value, str) and not value.strip())


def _trim_grid(grid: list[list[Any]]) -> list[list[Any]]:
    """Drop fully blank rows at the edges and fully blank columns anywhere."""
    rows = [r for r in grid if not all(_is_blank(c) for c in r)]
    if not rows:
        return []
    width = max(len(r) for r in rows)
    keep = [i for i in range(width) if not all(_is_blank(_at(r, i)) for r in rows)]
    return [[_at(r, i) for i in keep] for r in rows]


def _at(row: list[Any], index: int) -> Any:
    """Read ``row[index]`` treating a short row as blank-padded."""
    return row[index] if index < len(row) else None


def _detect_header_row(grid: list[list[Any]]) -> int | None:
    """Find the most likely header row, or ``None`` when nothing looks like one.

    A header is a row that is mostly populated, mostly textual, and followed by
    at least one row of data. Report titles fail the first test (one cell in a
    wide row); data rows fail the second.
    """
    width = max((len(r) for r in grid), default=0)
    if not width:
        return None

    best: tuple[float, int] | None = None
    limit = min(len(grid) - 1, _HEADER_SEARCH_ROWS)
    for i in range(limit):
        row = grid[i]
        filled = [c for c in row if not _is_blank(c)]
        if len(filled) < 2:
            continue
        fill_ratio = len(filled) / width
        text_ratio = sum(1 for c in filled if isinstance(c, str)) / len(filled)
        if fill_ratio < 0.6 or text_ratio < 0.6:
            continue
        # A header should not look like the rows beneath it. When the next row
        # is also all text this may be a two-line header or a text column, so
        # prefer the later row by scoring it slightly higher.
        following = grid[i + 1]
        following_filled = [c for c in following if not _is_blank(c)]
        if not following_filled:
            continue
        following_text = sum(1 for c in following_filled if isinstance(c, str)) / len(
            following_filled
        )
        score = fill_ratio + text_ratio - following_text
        if best is None or score > best[0]:
            best = (score, i)
    return best[1] if best else None


# Share of body rows that may deviate from the sheet's dominant column shape
# before the sheet is treated as messy.
_SHAPE_DEVIATION_LIMIT = 0.05


def _row_shape(row: list[Any], width: int) -> tuple[bool, ...]:
    """Which columns of a row hold a value -- the row's structural signature."""
    return tuple(not _is_blank(_at(row, i)) for i in range(width))


def _looks_messy(grid: list[list[Any]], header_row: int | None) -> bool:
    """Decide whether a sheet needs LLM help to locate its table.

    Merged header cells, stacked titles, subtotal lines and trailing note
    blocks all disturb the *shape* of the grid, so the grid alone is signal
    enough -- no need for ``worksheet.merged_cells``, which read-only mode does
    not expose anyway.

    The main test compares every body row against the sheet's dominant column
    signature. Real data rows fill the same columns as each other; a subtotal
    row leaves the middle columns blank and a "Notes:" line fills only the
    first, so both stand out without needing to match on language-specific
    keywords like "total" or "yhteensä".
    """
    if header_row is None:
        return True
    width = max((len(r) for r in grid), default=0)
    if not width:
        return True

    header = grid[header_row]
    filled = [c for c in header if not _is_blank(c)]
    # Merged headers leave gaps: a run of values then blanks across the row.
    if len(filled) < width:
        return True
    labels = [str(c).strip().lower() for c in filled]
    if len(set(labels)) != len(labels):
        return True

    body = grid[header_row + 1 :]
    if not body:
        return True

    shapes: dict[tuple[bool, ...], int] = {}
    for row in body:
        shapes[_row_shape(row, width)] = shapes.get(_row_shape(row, width), 0) + 1
    dominant = max(shapes.values())
    return (len(body) - dominant) / len(body) > _SHAPE_DEVIATION_LIMIT


def _resolve_layout(
    grid: list[list[Any]],
    *,
    sheet_label: str,
    layout_inference: str,
    infer_layout: LayoutInferrer | None,
) -> tuple[SheetLayout, bool]:
    """Pick the layout for a sheet, asking the LLM only when it is worth it.

    Returns:
        The layout, and whether the sheet looked messy -- the caller uses the
        latter to decide whether interleaved subtotal rows need removing.
    """
    header_row = _detect_header_row(grid)
    messy = _looks_messy(grid, header_row)
    heuristic = SheetLayout(
        header_row=header_row if header_row is not None else 0,
        data_start_row=(header_row if header_row is not None else 0) + 1,
        data_end_row=None,
        skip_columns=[],
        reasoning="Deterministic heuristic.",
    )

    if layout_inference == "never" or infer_layout is None:
        return heuristic, messy
    if layout_inference != "always" and not messy:
        return heuristic, messy

    try:
        inferred = infer_layout(sheet_label, render_grid(grid))
    except Exception as exc:
        logger.warning("Layout inference failed for %s, using heuristic: %s", sheet_label, exc)
        return heuristic, messy

    if not _layout_is_sane(inferred, grid):
        logger.warning("Layout inference returned an unusable result for %s", sheet_label)
        return heuristic, messy
    logger.info("Layout inferred for %s: %s", sheet_label, inferred.reasoning)
    return inferred, messy


def _layout_is_sane(layout: SheetLayout, grid: list[list[Any]]) -> bool:
    """Reject an inferred layout that points outside the sheet."""
    if not 0 <= layout.header_row < len(grid):
        return False
    if not layout.header_row < layout.data_start_row <= len(grid):
        return False
    if layout.data_end_row is not None and not (
        layout.data_start_row <= layout.data_end_row < len(grid)
    ):
        return False
    return True


def read_sheet_grid(path: str | Path, sheet_name: str | None = None) -> list[list[Any]]:
    """Read one worksheet as the trimmed cell grid the loader works from.

    Useful for inspecting what a spreadsheet actually contains before the
    header detection and clean-up run -- pair it with :func:`render_grid`.

    Args:
        path: Path to an Excel workbook.
        sheet_name: Which sheet to read; defaults to the first one.

    Returns:
        The grid with fully blank edge rows and blank columns removed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading Excel files requires 'openpyxl'. "
            "Install extras with 'pip install gaik[tabular-agent]'"
        ) from exc

    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        return _trim_grid(_read_grid(worksheet))
    finally:
        workbook.close()


def render_grid(grid: list[list[Any]]) -> str:
    """Render the top-left corner of a sheet as text for the LLM.

    Rows carry their 0-based index so the model can answer in the same
    coordinate system that :class:`SheetLayout` expects.
    """
    lines = []
    for i, row in enumerate(grid[:PREVIEW_ROWS]):
        cells = []
        for value in row[:PREVIEW_COLS]:
            text = "" if _is_blank(value) else str(value).strip().replace("\t", " ")
            cells.append(text[:40])
        lines.append(f"row {i}\t" + "\t".join(cells))
    if len(grid) > PREVIEW_ROWS:
        lines.append(f"... ({len(grid) - PREVIEW_ROWS} more row(s) not shown)")
    return "\n".join(lines)


def _apply_layout(
    grid: list[list[Any]], layout: SheetLayout, *, drop_outliers: bool = False
) -> tuple[list[str], list[list[Any]]]:
    """Slice a grid into ``(column_names, data_rows)`` using ``layout``.

    Args:
        grid: The trimmed cell grid.
        layout: Where the header and data live.
        drop_outliers: Also remove subtotal and section rows -- see
            :func:`_drop_shape_outliers`. Only meaningful for messy sheets.
    """
    skip = set(layout.skip_columns)
    width = len(grid[layout.header_row])
    keep = [i for i in range(width) if i not in skip]
    header = _name_columns([_at(grid[layout.header_row], i) for i in keep])

    end = layout.data_end_row + 1 if layout.data_end_row is not None else len(grid)
    rows: list[list[Any]] = []
    for raw in grid[layout.data_start_row : end]:
        row = [_at(raw, i) for i in keep]
        if all(_is_blank(c) for c in row):
            continue
        rows.append(row[: len(header)] + [None] * (len(header) - len(row)))

    if drop_outliers:
        rows = _drop_shape_outliers(rows)
    return header, rows


# A row filling at most this share of a normal row's columns is a section
# header or a note, whatever its columns happen to be.
_SPARSE_FILL_RATIO = 0.5
# Below this, no single row shape is "normal" enough to judge the others by.
_DOMINANT_SHARE_MIN = 0.6


def _numeric_columns(rows: list[list[Any]], dominant: tuple[bool, ...]) -> set[int]:
    """Indices of columns that hold numbers in the sheet's normal data rows."""
    width = len(dominant)
    numeric: set[int] = set()
    for i in range(width):
        if not dominant[i]:
            continue
        values = [_at(r, i) for r in rows if _row_shape(r, width) == dominant]
        values = [v for v in values if not _is_blank(v)]
        if values and all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in values):
            numeric.add(i)
    return numeric


def _drop_shape_outliers(rows: list[list[Any]]) -> list[list[Any]]:
    """Remove subtotal, section and note rows interleaved with the real data.

    ``SheetLayout`` can only cut a contiguous range, so a report with
    ``North subtotal`` after every region keeps those rows no matter how good
    the inference was. Two structural signatures catch them without matching on
    words like "total" or "yhteensä", which would only work in one language:

    * a **subtotal** blanks the descriptive columns but keeps the numbers, so
      the only columns it drops relative to a normal row are non-numeric ones;
    * a **note or section header** fills barely any columns at all.

    Deliberately conservative: it gives up when no shape dominates, never drops
    a row that fills *more* columns than the norm, and keeps rows that are
    missing a numeric value -- those read as real observations with a gap. The
    residual ambiguity is a data row missing one of its descriptive fields,
    which is structurally indistinguishable from a subtotal; see the README's
    known limitations. Applied only to sheets already flagged messy, so tidy
    exports are untouched.
    """
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    shapes: dict[tuple[bool, ...], int] = {}
    for row in rows:
        shape = _row_shape(row, width)
        shapes[shape] = shapes.get(shape, 0) + 1
    dominant, count = max(shapes.items(), key=lambda kv: kv[1])
    if count / len(rows) < _DOMINANT_SHARE_MIN:
        return rows

    dominant_filled = sum(dominant)
    if not dominant_filled:
        return rows
    numeric = _numeric_columns(rows, dominant)

    kept: list[list[Any]] = []
    for row in rows:
        shape = _row_shape(row, width)
        if shape != dominant and all(d for s, d in zip(shape, dominant) if s):
            missing = {i for i in range(width) if dominant[i] and not shape[i]}
            is_sparse = sum(shape) <= _SPARSE_FILL_RATIO * dominant_filled
            drops_only_labels = bool(missing) and not (missing & numeric)
            if is_sparse or drops_only_labels:
                continue
        kept.append(row)
    dropped = len(rows) - len(kept)
    if dropped:
        logger.info("Dropped %d subtotal/section row(s) from a messy sheet", dropped)
    return kept if kept else rows


def _name_columns(cells: list[Any]) -> list[str]:
    """Give every column a unique, non-empty, SQL-safe name."""
    names: list[str] = []
    taken: set[str] = set()
    for i, cell in enumerate(cells):
        fallback = f"column_{i + 1}"
        base = fallback if _is_blank(cell) else sanitize_identifier(str(cell), fallback=fallback)
        name = _unique(base, taken)
        taken.add(name)
        names.append(name)
    return names


def _create_from_rows(con: Any, table_name: str, header: list[str], rows: list[list[Any]]) -> None:
    """Create a table from in-memory rows via a temporary CSV.

    Routing Excel through DuckDB's CSV reader rather than inserting typed values
    directly means both input paths get the same type inference, so a column of
    dates behaves identically whether it arrived as .csv or .xlsx.
    """
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".csv", delete=False, newline="", encoding="utf-8"
    )
    try:
        writer = csv.writer(tmp)
        writer.writerow(header)
        for row in rows:
            writer.writerow(["" if _is_blank(c) else c for c in row])
        tmp.close()
        posix = Path(tmp.name).as_posix()
        con.execute(
            f'CREATE TABLE "{table_name}" AS SELECT * FROM '
            f"read_csv('{_escape(posix)}', auto_detect=true, sample_size=-1)"
        )
    except Exception as exc:
        raise TableLoadError(f"Could not build table '{table_name}': {exc}") from exc
    finally:
        Path(tmp.name).unlink(missing_ok=True)
    repair_european_decimals(con, table_name)


def lock_down(con: Any) -> None:
    """Close the connection's access to everything outside the loaded tables.

    Called once, after loading and before any LLM-generated SQL runs. Both
    settings are irreversible for the life of the connection: DuckDB refuses to
    re-enable external access once the configuration is locked.
    """
    con.execute("SET enable_external_access = false")
    con.execute("SET lock_configuration = true")
