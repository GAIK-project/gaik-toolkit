"""Unit tests for tabular_agent file loading and spreadsheet clean-up.

These build their own fixtures in ``tmp_path`` and never call an LLM: layout
inference is exercised through an injected callback, so the deterministic tiers
and the LLM tier can both be tested offline.
"""

import pytest

pytest.importorskip("duckdb")
pytest.importorskip("openpyxl")

import duckdb  # noqa: E402
from gaik.software_components.tabular_agent.loader import (  # noqa: E402
    TableLoadError,
    _detect_header_row,
    _drop_shape_outliers,
    _looks_messy,
    _read_grid,
    _trim_grid,
    load_sources,
    normalize_sources,
    sanitize_identifier,
)
from gaik.software_components.tabular_agent.models import SheetLayout  # noqa: E402
from openpyxl import Workbook  # noqa: E402

CLEAN_ROWS = [
    ("North", "Q1", 120, 14400.0),
    ("South", "Q1", 90, 10800.0),
    ("East", "Q1", 210, 25200.0),
    ("West", "Q1", 60, 7200.0),
]


@pytest.fixture
def con():
    connection = duckdb.connect(":memory:")
    yield connection
    connection.close()


def _load(con, path, **kwargs):
    """Load one file and return ``{table_name: source}``."""
    return load_sources(con, normalize_sources(path), **kwargs)


# ---------------------------------------------------------------- identifiers


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("Sales Report", "sales_report"),
        ("Revenue (EUR)", "revenue_eur"),
        ("2026 totals", "t_2026_totals"),
        ("  ", "table"),
        ("Myynti/Alue", "myynti_alue"),
    ],
)
def test_sanitize_identifier(raw, expected):
    assert sanitize_identifier(raw) == expected


# ------------------------------------------------------------------ CSV paths


def test_loads_plain_csv(con, tmp_path):
    path = tmp_path / "sales.csv"
    path.write_text(
        "region,quarter,units,revenue\n"
        + "\n".join(f"{r},{q},{u},{v}" for r, q, u, v in CLEAN_ROWS)
        + "\n",
        encoding="utf-8",
    )
    tables = _load(con, path)
    assert list(tables) == ["sales"]
    assert con.execute("SELECT count(*) FROM sales").fetchone()[0] == 4


def test_loads_semicolon_csv_with_bom_and_comma_decimals(con, tmp_path):
    """Nordic exports: ';' separator, ',' decimal mark, UTF-8 BOM."""
    path = tmp_path / "myynti.csv"
    path.write_text(
        "region;quarter;units;revenue\n"
        + "\n".join(f"{r};{q};{u};{v:.2f}".replace(".", ",") for r, q, u, v in CLEAN_ROWS)
        + "\n",
        encoding="utf-8-sig",
    )
    _load(con, path)
    types = {r[0]: r[1] for r in con.execute("DESCRIBE myynti").fetchall()}
    assert types["revenue"] == "DOUBLE", "comma-decimals must be usable in SUM()"
    assert con.execute("SELECT sum(revenue) FROM myynti").fetchone()[0] == pytest.approx(57600.0)
    assert con.execute("SELECT count(*) FROM myynti").fetchone()[0] == 4


def test_free_text_with_commas_is_not_converted_to_numbers(con, tmp_path):
    """The decimal repair must only fire when every value is a number."""
    path = tmp_path / "notes.csv"
    path.write_text("id;comment\n1;hello, world\n2;3,5 was the score\n", encoding="utf-8")
    _load(con, path)
    types = {r[0]: r[1] for r in con.execute("DESCRIBE notes").fetchall()}
    assert types["comment"].startswith("VARCHAR")


def test_rejects_unsupported_extension(tmp_path):
    path = tmp_path / "notes.pdf"
    path.write_bytes(b"%PDF-1.4")
    with pytest.raises(TableLoadError, match="Unsupported file type"):
        normalize_sources(path)


def test_rejects_missing_file(tmp_path):
    with pytest.raises(TableLoadError, match="File not found"):
        normalize_sources(tmp_path / "nope.csv")


def test_mapping_source_names_the_table(con, tmp_path):
    path = tmp_path / "raw_export_2026.csv"
    path.write_text("a,b\n1,2\n", encoding="utf-8")
    tables = load_sources(con, normalize_sources({"myynti": path}))
    assert list(tables) == ["myynti"]


def test_two_files_with_the_same_stem_get_unique_names(con, tmp_path):
    first = tmp_path / "a" / "data.csv"
    second = tmp_path / "b" / "data.csv"
    for path in (first, second):
        path.parent.mkdir(parents=True)
        path.write_text("x\n1\n", encoding="utf-8")
    tables = load_sources(con, normalize_sources([first, second]))
    assert list(tables) == ["data", "data_2"]


# ---------------------------------------------------------------- Excel paths


def _write_clean_xlsx(path, sheets):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(list(row))
    workbook.save(path)


def test_loads_one_table_per_sheet(con, tmp_path):
    path = tmp_path / "book.xlsx"
    _write_clean_xlsx(
        path,
        {
            "Sales": [("region", "units")] + [(r, u) for r, _, u, _ in CLEAN_ROWS],
            "Headcount": [("region", "employees"), ("North", 12), ("South", 9)],
        },
    )
    tables = _load(con, path)
    assert set(tables) == {"sales", "headcount"}
    assert con.execute("SELECT count(*) FROM sales").fetchone()[0] == 4
    assert con.execute("SELECT count(*) FROM headcount").fetchone()[0] == 2


def test_single_sheet_workbook_is_named_after_the_file(con, tmp_path):
    path = tmp_path / "quarterly.xlsx"
    _write_clean_xlsx(path, {"Sheet1": [("a", "b"), (1, 2)]})
    assert list(_load(con, path)) == ["quarterly"]


def test_blank_leading_rows_and_columns_are_dropped(con, tmp_path):
    path = tmp_path / "padded.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([])
    sheet.append([None, None, "region", "units"])
    sheet.append([None, None, "North", 120])
    sheet.append([None, None, "South", 90])
    workbook.save(path)
    _load(con, path)
    columns = [r[0] for r in con.execute("DESCRIBE padded").fetchall()]
    assert columns == ["region", "units"]
    assert con.execute("SELECT count(*) FROM padded").fetchone()[0] == 2


def test_unnamed_columns_get_generated_names(con, tmp_path):
    path = tmp_path / "partial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["region", None, "units"])
    sheet.append(["North", "x", 120])
    sheet.append(["South", "y", 90])
    workbook.save(path)
    _load(con, path)
    columns = [r[0] for r in con.execute("DESCRIBE partial").fetchall()]
    assert columns[0] == "region"
    assert columns[1].startswith("column_")


def test_empty_sheets_are_skipped(con, tmp_path):
    path = tmp_path / "mixed.xlsx"
    _write_clean_xlsx(path, {"Empty": [], "Real": [("a",), (1,)]})
    assert list(_load(con, path)) == ["real"]


def test_workbook_with_no_usable_sheet_raises(con, tmp_path):
    path = tmp_path / "blank.xlsx"
    _write_clean_xlsx(path, {"Empty": []})
    with pytest.raises(TableLoadError, match="No usable sheets"):
        _load(con, path)


# ------------------------------------------------------- messy-sheet handling


def _messy_workbook(path):
    """A human-facing report: title rows, subtotals, and a trailing note block."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(["ACME -- Quarterly Sales"])
    sheet.append(["Figures in EUR"])
    sheet.append([])
    sheet.append(["Region", "Quarter", "Units", "Revenue"])
    for region, _, units, revenue in CLEAN_ROWS:
        for quarter in ("Q1", "Q2", "Q3"):
            sheet.append([region, quarter, units, revenue])
        sheet.append([f"{region} subtotal", None, units * 3, revenue * 3])
    sheet.append([])
    sheet.append(["Notes:"])
    sheet.append(["Unaudited."])
    workbook.save(path)


def _grid(path, sheet_name):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return _trim_grid(_read_grid(workbook[sheet_name]))
    finally:
        workbook.close()


def test_header_row_found_below_title_rows(tmp_path):
    path = tmp_path / "messy.xlsx"
    _messy_workbook(path)
    grid = _grid(path, "Report")
    # Row 2 after the blank spacer row is trimmed away.
    assert _detect_header_row(grid) == 2


def test_messy_sheet_is_flagged_and_clean_sheet_is_not(tmp_path):
    messy = tmp_path / "messy.xlsx"
    _messy_workbook(messy)
    messy_grid = _grid(messy, "Report")
    assert _looks_messy(messy_grid, _detect_header_row(messy_grid)) is True

    clean = tmp_path / "clean.xlsx"
    _write_clean_xlsx(
        clean, {"Sales": [("region", "units")] + [(r, u) for r, _, u, _ in CLEAN_ROWS]}
    )
    clean_grid = _grid(clean, "Sales")
    assert _looks_messy(clean_grid, _detect_header_row(clean_grid)) is False


def test_subtotal_rows_are_dropped_structurally():
    """A subtotal blanks the label columns but keeps the numbers."""
    rows = []
    for region in ("North", "South", "East"):
        rows += [
            [region, "Q1", 120, 14400.0],
            [region, "Q2", 90, 10800.0],
            [region, "Q3", 60, 7200.0],
        ]
        rows.append([f"{region} subtotal", None, 270, 32400.0])
    kept = _drop_shape_outliers(rows)
    assert len(kept) == 9
    assert all(row[1] is not None for row in kept)


def test_note_rows_are_dropped():
    """A trailing note fills almost nothing, whatever its columns are."""
    rows = [["North", "Q1", 120, 14400.0] for _ in range(9)]
    rows += [["Notes:", None, None, None], ["Unaudited.", None, None, None]]
    assert len(_drop_shape_outliers(rows)) == 9


def test_rows_missing_a_numeric_value_are_kept():
    """A real observation with a gap in a measure is not a subtotal."""
    rows = [["North", "Q1", 120, 14400.0] for _ in range(9)]
    rows.append(["South", "Q1", 90, None])
    assert len(_drop_shape_outliers(rows)) == 10


def test_outlier_filter_gives_up_without_a_dominant_shape():
    rows = [["a", None, None], [None, "b", None], [None, None, "c"]]
    assert _drop_shape_outliers(rows) == rows


def test_layout_inference_callback_trims_trailing_notes(con, tmp_path):
    """The LLM tier is exercised through an injected callback -- no API call."""
    path = tmp_path / "messy.xlsx"
    _messy_workbook(path)
    calls = []

    def fake_infer(sheet_label, grid_text):
        calls.append(sheet_label)
        assert "row 0" in grid_text
        return SheetLayout(
            header_row=2,
            data_start_row=3,
            data_end_row=18,  # stop before the notes block
            skip_columns=[],
            reasoning="test",
        )

    _load(con, path, layout_inference="auto", infer_layout=fake_infer)
    assert calls, "a messy sheet must trigger layout inference"
    regions = [r[0] for r in con.execute("SELECT region FROM messy").fetchall()]
    assert "Notes:" not in regions
    assert not any("subtotal" in str(r) for r in regions)


def test_clean_sheet_does_not_trigger_layout_inference(con, tmp_path):
    """Tidy files must cost zero extra tokens."""
    path = tmp_path / "clean.xlsx"
    _write_clean_xlsx(
        path, {"Sales": [("region", "units")] + [(r, u) for r, _, u, _ in CLEAN_ROWS]}
    )
    calls = []

    def fake_infer(sheet_label, grid_text):
        calls.append(sheet_label)
        raise AssertionError("should not be called for a clean sheet")

    _load(con, path, layout_inference="auto", infer_layout=fake_infer)
    assert calls == []


def test_never_mode_skips_inference_even_when_messy(con, tmp_path):
    path = tmp_path / "messy.xlsx"
    _messy_workbook(path)

    def fake_infer(sheet_label, grid_text):
        raise AssertionError("should not be called in 'never' mode")

    _load(con, path, layout_inference="never", infer_layout=fake_infer)
    assert con.execute("SELECT count(*) FROM messy").fetchone()[0] > 0


def test_unusable_inferred_layout_falls_back_to_heuristic(con, tmp_path):
    path = tmp_path / "messy.xlsx"
    _messy_workbook(path)

    def bad_infer(sheet_label, grid_text):
        return SheetLayout(
            header_row=999,
            data_start_row=1000,
            data_end_row=None,
            skip_columns=[],
            reasoning="nonsense",
        )

    _load(con, path, layout_inference="auto", infer_layout=bad_infer)
    assert con.execute("SELECT count(*) FROM messy").fetchone()[0] > 0


def test_inference_failure_falls_back_to_heuristic(con, tmp_path):
    path = tmp_path / "messy.xlsx"
    _messy_workbook(path)

    def failing_infer(sheet_label, grid_text):
        raise RuntimeError("LLM unavailable")

    _load(con, path, layout_inference="auto", infer_layout=failing_infer)
    assert con.execute("SELECT count(*) FROM messy").fetchone()[0] > 0
