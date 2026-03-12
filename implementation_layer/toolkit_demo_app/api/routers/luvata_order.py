"""
Luvata Order Processing API Router
Handles ABB purchase order processing with BOM matching and pricing calculations.
Demonstrates GAIK toolkit's extraction capabilities for real-world manufacturing workflows.
"""
import csv
import os
import io
import pandas as pd
from io import BytesIO
import logging
import re
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Annotated
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from fpdf import FPDF
from openpyxl import load_workbook
from pydantic import BaseModel, Field
try:
    from utils.config import get_api_config
    from utils.sse import sse_event
except ImportError:
    from api.utils.config import get_api_config
    from api.utils.sse import sse_event
from gaik.software_components.extractor import DataExtractor, SchemaGenerator, ExtractionRequirements, FieldSpec
from gaik.software_components.extractor.schema import print_pydantic_schema
from gaik.software_components.parsers import PyMuPDFParser
from gaik.software_components.parsers.docling_api_client import DoclingApiClientParser
import importlib.util
import json
from contextlib import redirect_stdout
logger = logging.getLogger(__name__)
router = APIRouter(prefix="/luvata-order", tags=["luvata-order"])
# Schema directory
SCHEMA_DIR = Path(__file__).parent.parent / "schemas"
SCHEMA_DIR.mkdir(exist_ok=True)
DEBUG_MARKDOWN_DIR = Path(__file__).parent.parent / "debug_markdown"
DEBUG_MARKDOWN_DIR.mkdir(exist_ok=True)
def _clean_schema_dump(raw_dump: str) -> str:
    """Strip header/footer lines from print_pydantic_schema output."""
    lines = raw_dump.splitlines()
    start_idx = 0
    for i, line in enumerate(lines):
        if line.startswith("class "):
            start_idx = i
            break
    body = lines[start_idx:]
    while body and (set(body[-1].strip()) == {"="} or not body[-1].strip()):
        body.pop()
    return "\n".join(body).strip()
def _sanitize_schema_code(schema_code: str) -> str:
    """
    Normalize generated schema code so cached modules are self-contained.
    `print_pydantic_schema()` may emit fully qualified references back into
    `gaik.software_components.extractor.schema`. Those names are not available
    inside the cached module unless the whole package path is imported, and the
    referenced classes are already emitted in the same file anyway.
    """
    schema_code = schema_code.replace(
        "gaik.software_components.extractor.schema.", ""
    )
    return schema_code
def save_schema_to_python(model: type, path: Path) -> None:
    """Dump the generated Pydantic model into a valid Python file."""
    buffer = io.StringIO()
    with redirect_stdout(buffer):
        print_pydantic_schema(model, title="Generated Schema")
    schema_code = _sanitize_schema_code(_clean_schema_dump(buffer.getvalue()))
    template = f'''"""
Auto-generated schema module (do not edit manually).
"""
import decimal
from decimal import Decimal
from typing import List, Literal, Optional
from pydantic import BaseModel, Field, ConfigDict
{schema_code}
'''
    path.write_text(template, encoding="utf-8")
    logger.info(f"Schema saved to {path}")
def save_requirements(requirements: ExtractionRequirements, model_name: str, path: Path) -> None:
    """Save extraction requirements to JSON."""
    payload = {
        "model_name": model_name,
        "requirements": requirements.model_dump(),
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    logger.info(f"Requirements saved to {path}")
def load_saved_schema(path: Path, model_name: str):
    """Load the previously saved schema module and return the model class."""
    source = path.read_text(encoding="utf-8")
    sanitized = _sanitize_schema_code(source)
    if sanitized != source:
        path.write_text(sanitized, encoding="utf-8")
        logger.info(f"Sanitized cached schema module: {path}")
    spec = importlib.util.spec_from_file_location("saved_schema", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    try:
        spec.loader.exec_module(module)
    except Exception:
        logger.exception(f"Failed to load cached schema module: {path}")
        raise
    model_cls = getattr(module, model_name)
    return model_cls
def load_saved_requirements(path: Path) -> tuple[str, ExtractionRequirements]:
    """Load extraction requirements from JSON."""
    data = json.loads(path.read_text(encoding="utf-8"))
    model_name = data["model_name"]
    requirements = ExtractionRequirements(**data["requirements"])
    return model_name, requirements
# In-memory storage for generated PDFs (temporary, with cleanup)
pdf_storage: dict[str, bytes] = {}
# ============================================================================
# Pydantic Models
# ============================================================================
class POItem(BaseModel):
    """Purchase Order line item"""
    material: str = Field(description="Material number/code")
    description: str = Field(description="Item description")
    quantity: int = Field(description="Order quantity")
    delivery_date: str | None = Field(None, description="Requested delivery date")
class PurchaseOrder(BaseModel):
    """Purchase Order data"""
    po_number: str = Field(description="Purchase order number")
    customer: str = Field(description="Customer name")
    customer_address: str | None = Field(default=None, description="Customer/buyer address block")
    delivery_address: str | None = Field(default=None, description="Delivery address block")
    invoicing_address: str | None = Field(default=None, description="Invoicing address block")
    items: list[POItem] = Field(description="Order line items")
class BOMData(BaseModel):
    """Bill of Materials data"""
    material_id: str = Field(description="Material ID (e.g., AL-001)")
    type_designation: str = Field(description="Product type designation")
    dimensions: str = Field(description="Product dimensions")
    material_grade: str = Field(description="Material grade/standard")
    cutting_required: bool = Field(default=False, description="Cutting service required")
    testing_required: bool = Field(default=False, description="Testing service required")
    certificates_required: bool = Field(default=False, description="Certificates required")
class PricingRow(BaseModel):
    """Pricing table row"""
    material_id: str | None = Field(default=None, description="Material/item identifier from the pricing table")
    type_designation: str = Field(description="Product type designation")
    unit_price: float = Field(description="Base unit price (USD)")
    cutting_fee: float = Field(default=0.0, description="Cutting fee (USD)")
    testing_fee: float = Field(default=0.0, description="Testing fee (USD)")
    cert_fee: float = Field(default=0.0, description="Certificate fee (USD)")
class EnrichedItem(BaseModel):
    """Enriched order item with calculated pricing"""
    material: str
    description: str
    type_designation: str | None = None
    quantity: int
    unit_price: float | None = None
    material_subtotal: float | None = None
    cutting_fee: float = 0.0
    testing_fee: float = 0.0
    cert_fee: float = 0.0
    total_fees: float = 0.0
    line_total: float | None = None
    delivery_date: str | None = None
    bom_match: bool = False
    price_match: bool = False
    error: str | None = None
class OrderSummary(BaseModel):
    """Order summary totals"""
    total_items: int
    total_quantity: int
    material_subtotal: float
    total_fees: float
    grand_total: float
class ProcessOrderResponse(BaseModel):
    """Response from order processing"""
    success: bool
    po_number: str | None = None
    customer: str | None = None
    items: list[EnrichedItem] = []
    summary: OrderSummary | None = None
    errors: list[str] = []
    warnings: list[str] = []
    pdf_job_id: str | None = None
# ============================================================================
# Helper Functions
# ============================================================================
def normalize_type_designation(type_des: str) -> str:
    """
    Normalize type designation for matching.
    Example: "BKMJ 50X10/7,00" → "bkmj50x10/7"
    """
    normalized = type_des.lower()
    # Remove spaces, commas, parentheses
    normalized = normalized.replace(" ", "").replace(",", "").replace("(", "").replace(")", "")
    # Remove trailing zeros after slash
    if "/" in normalized:
        parts = normalized.split("/")
        if len(parts) == 2 and parts[1].replace("0", "").replace(".", "") == "":
            normalized = parts[0] + "/" + parts[1].rstrip("0").rstrip(".")
    return normalized
def _save_parsed_markdown(prefix: str, filename: str | None, markdown: str) -> Path:
    """Persist parsed markdown for temporary debugging and inspection."""
    source_name = Path(filename or prefix).stem
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", source_name).strip("_") or prefix
    output_path = DEBUG_MARKDOWN_DIR / f"{prefix}_{safe_name}.md"
    output_path.write_text(markdown, encoding="utf-8")
    logger.info(f"Saved parsed markdown to {output_path}")
    return output_path
def _parse_document_markdown(file_path: str, original_filename: str | None, prefix: str) -> str:
    """Parse a document via remote Docling client when configured, otherwise fallback locally."""
    api_base = os.getenv("DOCLING_API_BASE") or os.getenv("API_BASE")
    password = os.getenv("DOCLING_API_PASSWORD") or os.getenv("PASSWORD")
    if api_base and password:
        try:
            remote_parser = DoclingApiClientParser(api_base=api_base, password=password)
            result = remote_parser.parse_document(file_path)
            markdown = (result.get("parsed_markdown") or "").strip()
            if markdown:
                logger.info("Parsed %s via Docling API client", original_filename or file_path)
                _save_parsed_markdown(prefix, original_filename, markdown)
                return markdown
            logger.warning(
                "Docling API client returned empty markdown for %s; falling back to PyMuPDFParser",
                original_filename or file_path,
            )
        except Exception as exc:
            logger.warning(
                "Docling API client unavailable for %s; falling back to PyMuPDFParser: %s",
                original_filename or file_path,
                exc,
            )
    local_parser = PyMuPDFParser()
    markdown = local_parser.parse_pdf(file_path, use_markdown=True)
    _save_parsed_markdown(prefix, original_filename, markdown)
    return markdown
def _parse_quantity(raw_quantity: int | str) -> int:
    """Parse quantity robustly (supports EU/US separators)."""
    if isinstance(raw_quantity, int):
        return raw_quantity
    s = str(raw_quantity).strip()
    if not s:
        return 0
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "." in s:
        parts = s.split(".")
        if (
            len(parts) == 2
            and len(parts[1]) == 3
            and parts[1].isdigit()
            and len(parts[0].lstrip("-")) <= 3
        ):
            s = s.replace(".", "")
    elif "," in s:
        parts = s.split(",")
        if (
            len(parts) == 2
            and len(parts[1]) == 3
            and parts[1].isdigit()
            and len(parts[0].lstrip("-")) <= 3
        ):
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    return int(float(s))
def _extract_fee_flags_from_bom_text(text: str) -> dict[str, bool]:
    """
    Extract BOM fee flags from raw text.
    The BOM PDFs often print fee labels together and place their values later in
    the same order. We therefore support both patterns:
    1. value immediately after the label
    2. grouped values appearing later in the section
    """
    lines = [line.strip() for line in text.splitlines()]
    label_patterns = {
        "testing_required": re.compile(r"^testing required\s*:?$", re.IGNORECASE),
        "cutting_required": re.compile(r"^cutting required\s*:?$", re.IGNORECASE),
        "certificates_required": re.compile(
            r"^(cert\.? required|certificates required)\s*:?$", re.IGNORECASE
        ),
    }
    def parse_value(value: str) -> bool | None:
        normalized = " ".join(value.lower().split())
        if not normalized:
            return None
        if normalized in {"no", "not required", "none", "n/a"}:
            return False
        if any(token in normalized for token in ["yes", "1 lot", "1 certificate", "certificate", "lot"]):
            return True
        if re.search(r"\d+\s*(lot|lots|certificate|certificates)", normalized):
            return True
        return None
    label_positions: dict[str, list[int]] = {key: [] for key in label_patterns}
    for idx, line in enumerate(lines):
        for key, pattern in label_patterns.items():
            if pattern.match(line):
                label_positions[key].append(idx)
    values: dict[str, bool | None] = {key: None for key in label_patterns}
    # First pass: detect values immediately after each label.
    for key, positions in label_positions.items():
        for pos in positions:
            for next_idx in range(pos + 1, min(pos + 5, len(lines))):
                candidate = parse_value(lines[next_idx])
                if candidate is not None:
                    values[key] = candidate
                    break
            if values[key] is not None:
                break
    unresolved = [key for key, value in values.items() if value is None and label_positions[key]]
    if unresolved:
        ordered_labels = sorted(
            [(positions[0], key) for key, positions in label_positions.items() if positions and values[key] is None],
            key=lambda item: item[0],
        )
        start = min(index for index, _ in ordered_labels)
        candidate_values: list[bool] = []
        for line in lines[start:]:
            candidate = parse_value(line)
            if candidate is not None:
                candidate_values.append(candidate)
        for (_, key), candidate in zip(ordered_labels, candidate_values):
            values[key] = candidate
    return {
        "testing_required": bool(values["testing_required"]),
        "cutting_required": bool(values["cutting_required"]),
        "certificates_required": bool(values["certificates_required"]),
    }
def _extract_po_address_block(text: str) -> tuple[str, str | None]:
    """Extract the buyer header block from parsed PO markdown."""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    stop_markers = (
        "Authorized By:",
        "This purchase order",
        "Special Instructions:",
        "PURCHASE ORDER",
        "SUPPLIER INFORMATION",
    )
    header_lines: list[str] = []
    for line in lines:
        if any(marker in line for marker in stop_markers):
            break
        header_lines.append(line)
        if len(header_lines) >= 3:
            break
    customer_name = header_lines[0] if header_lines else ""
    address_block = "\n".join(header_lines[1:]).strip() or None
    return customer_name, address_block
def match_material_to_bom(material: str, boms: list[BOMData]) -> BOMData | None:
    """
    Match PO material number to BOM material ID.
    Simple exact match on material_id field.
    """
    material_upper = material.upper().strip()
    for bom in boms:
        if bom.material_id.upper().strip() == material_upper:
            return bom
    return None
def match_type_to_pricing(
    type_designation: str, pricing_rows: list[PricingRow]
) -> PricingRow | None:
    """
    Match type designation to pricing table row.
    Supports both generic and ABB-style type strings.
    """
    normalized_type = normalize_type_designation(type_designation)
    for row in pricing_rows:
        if normalize_type_designation(row.type_designation) == normalized_type:
            return row
    for row in pricing_rows:
        normalized_row = normalize_type_designation(row.type_designation)
        if normalized_type in normalized_row or normalized_row in normalized_type:
            return row
    return None
def match_pricing(
    material: str, type_designation: str, pricing_rows: list[PricingRow]
) -> PricingRow | None:
    """Match pricing primarily by material/item id, then by type designation."""
    material_upper = material.upper().strip()
    for row in pricing_rows:
        if row.material_id and row.material_id.upper().strip() == material_upper:
            return row
    return match_type_to_pricing(type_designation, pricing_rows)
def calculate_simple_pricing(
    item: POItem, bom: BOMData, pricing: PricingRow
) -> EnrichedItem:
    """
    Calculate simple unit-based pricing:
    material_subtotal = quantity × unit_price
    fees = cutting_fee + testing_fee + cert_fee (only if required per BOM)
    line_total = material_subtotal + fees
    """
    quantity = _parse_quantity(item.quantity)
    # 1. Material cost
    material_subtotal = quantity * pricing.unit_price
    # 2. Optional fees (only if BOM indicates they're required)
    cutting_fee = pricing.cutting_fee if bom.cutting_required else 0.0
    testing_fee = pricing.testing_fee if bom.testing_required else 0.0
    cert_fee = pricing.cert_fee if bom.certificates_required else 0.0
    total_fees = cutting_fee + testing_fee + cert_fee
    # 3. Line total
    line_total = material_subtotal + total_fees
    return EnrichedItem(
        material=item.material,
        description=item.description,
        type_designation=bom.type_designation,
        quantity=quantity,
        unit_price=round(pricing.unit_price, 2),
        material_subtotal=round(material_subtotal, 2),
        cutting_fee=round(cutting_fee, 2),
        testing_fee=round(testing_fee, 2),
        cert_fee=round(cert_fee, 2),
        total_fees=round(total_fees, 2),
        line_total=round(line_total, 2),
        delivery_date=item.delivery_date,
        bom_match=True,
        price_match=True,
    )
def enrich_products(
    po: PurchaseOrder, boms: list[BOMData], pricing_rows: list[PricingRow]
) -> tuple[list[EnrichedItem], list[str]]:
    """
    Enrich PO items with BOM data and pricing calculations.
    Returns (enriched_items, errors)
    """
    enriched_items: list[EnrichedItem] = []
    errors: list[str] = []
    for item in po.items:
        # Match to BOM
        bom = match_material_to_bom(item.material, boms)
        if not bom:
            error_msg = f"No BOM found for material {item.material}"
            errors.append(error_msg)
            enriched_items.append(
                EnrichedItem(
                    material=item.material,
                    description=item.description,
                    quantity=item.quantity,
                    delivery_date=item.delivery_date,
                    error=error_msg,
                )
            )
            continue
        # Match to pricing
        pricing = match_pricing(item.material, bom.type_designation, pricing_rows)
        if not pricing:
            error_msg = (
                f"No pricing found for material '{item.material}' or type designation '{bom.type_designation}'"
            )
            errors.append(error_msg)
            enriched_items.append(
                EnrichedItem(
                    material=item.material,
                    description=item.description,
                    quantity=item.quantity,
                    type_designation=bom.type_designation,
                    delivery_date=item.delivery_date,
                    bom_match=True,
                    price_match=False,
                    error=error_msg,
                )
            )
            continue
        # Calculate pricing
        enriched_item = calculate_simple_pricing(item, bom, pricing)
        enriched_items.append(enriched_item)
    return enriched_items, errors
def calculate_summary(items: list[EnrichedItem]) -> OrderSummary:
    """Calculate order summary totals"""
    total_items = len(items)
    total_quantity = sum(item.quantity for item in items)
    material_subtotal = sum(item.material_subtotal or 0 for item in items)
    total_fees = sum(item.total_fees for item in items)
    grand_total = sum(item.line_total or 0 for item in items)
    return OrderSummary(
        total_items=total_items,
        total_quantity=total_quantity,
        material_subtotal=round(material_subtotal, 2),
        total_fees=round(total_fees, 2),
        grand_total=round(grand_total, 2),
    )
async def extract_po_data(po_file: UploadFile) -> PurchaseOrder:
    """Extract purchase order data from PDF using GAIK with schema caching"""
    pdf_bytes = await po_file.read()
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    try:
        text = _parse_document_markdown(tmp_path, po_file.filename, "po")
        # Get API config
        config = get_api_config()
        # Define schema paths
        schema_path = SCHEMA_DIR / "po_schema.py"
        requirements_path = SCHEMA_DIR / "po_requirements.json"
        user_requirements = (
            "Extract purchase order number, customer or buyer name, and line items. "
            "If available, also extract customer_address, delivery_address, and invoicing_address as text blocks. "
            "Each item should include material number, description, quantity as integer, and delivery date if available."
        )
        # Check if schema exists
        if schema_path.exists() and requirements_path.exists():
            logger.info("Loading existing PO schema...")
            model_name, requirements = load_saved_requirements(requirements_path)
            POModel = load_saved_schema(schema_path, model_name)
        else:
            logger.info("Generating new PO schema...")
            generator = SchemaGenerator(config=config)
            POModel = generator.generate_schema(user_requirements=user_requirements)
            
            # Save schema and requirements
            save_schema_to_python(POModel, schema_path)
            save_requirements(generator.item_requirements, POModel.__name__, requirements_path)
            requirements = generator.item_requirements
        # Extract data
        extractor = DataExtractor(config)
        result = extractor.extract(
            extraction_model=POModel,
            requirements=requirements,
            user_requirements=user_requirements,
            documents=[text],
        )
        if not result or not result[0]:
            raise ValueError("Failed to extract purchase order data")
        # Convert to PurchaseOrder object
        po_data = result[0]
        # Handle different field names from schema generator (line_items, items)
        line_items = po_data.get("line_items") or po_data.get("items", [])
        header_customer, header_address = _extract_po_address_block(text)
        # Extract PO number and customer from first item (duplicated in each item)
        po_number = po_data.get("purchase_order_number") or po_data.get("po_number") or ""
        customer = po_data.get("customer_name") or po_data.get("customer") or ""
        customer_address = (
            po_data.get("customer_address")
            or po_data.get("buyer_address")
            or po_data.get("address")
            or None
        )
        delivery_address = po_data.get("delivery_address") or None
        invoicing_address = po_data.get("invoicing_address") or po_data.get("invoice_address") or None
        if line_items:
            first_item = line_items[0]
            po_number = po_number or first_item.get("purchase_order_number") or first_item.get("po_number") or ""
            customer = customer or first_item.get("customer_name") or first_item.get("customer") or ""
            customer_address = customer_address or first_item.get("customer_address") or first_item.get("buyer_address")
            delivery_address = delivery_address or first_item.get("delivery_address")
            invoicing_address = invoicing_address or first_item.get("invoicing_address") or first_item.get("invoice_address")
        customer = customer or header_customer
        customer_address = customer_address or header_address
        delivery_address = delivery_address or customer_address
        invoicing_address = invoicing_address or customer_address
        normalized_items = []
        for item in line_items:
            material = item.get("material_number") or item.get("material", "")
            normalized_items.append(
                POItem(
                    material=material,
                    description=item.get("description", ""),
                    quantity=_parse_quantity(item.get("quantity", 0)),
                    delivery_date=item.get("delivery_date"),
                )
            )
        return PurchaseOrder(
            po_number=po_number,
            customer=customer,
            customer_address=customer_address,
            delivery_address=delivery_address,
            invoicing_address=invoicing_address,
            items=normalized_items,
        )
    finally:
        Path(tmp_path).unlink(missing_ok=True)
async def extract_bom_data(bom_file: UploadFile) -> BOMData:
    """Extract BOM data from PDF using GAIK with schema caching"""
    pdf_bytes = await bom_file.read()
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(pdf_bytes)
        tmp_path = tmp.name
    try:
        text = _parse_document_markdown(tmp_path, bom_file.filename, "bom")
        # Get API config
        config = get_api_config()
        # Define schema paths
        schema_path = SCHEMA_DIR / "bom_schema.py"
        requirements_path = SCHEMA_DIR / "bom_requirements.json"
        user_requirements = (
            "Extract material ID, product type designation, dimensions, and material grade. "
            "Also extract the boolean flags cutting_required, testing_required, and certificates_required. "
            "Use only the BOM fee labels and their values for these booleans: 'Testing Required:', "
            "'Cutting Required:', and 'Cert. Required:' or 'Certificates Required:'. "
            "The values may appear on later lines in the same order as the labels. "
            "Map '1 lot', 'Yes', or any explicit testing quantity to testing_required=True. "
            "Map '1 certificate', 'certificate required', 'Yes', or any explicit certificate quantity to "
            "certificates_required=True. "
            "Map 'No' to False for the corresponding label only. "
            "Do not infer fee flags from unrelated technical text, and do not mix testing, cutting, and "
            "certificate values with each other."
        )
        # Check if schema exists
        if schema_path.exists() and requirements_path.exists():
            logger.info("Loading existing BOM schema...")
            model_name, requirements = load_saved_requirements(requirements_path)
            BOMModel = load_saved_schema(schema_path, model_name)
        else:
            logger.info("Generating new BOM schema...")
            generator = SchemaGenerator(config=config)
            BOMModel = generator.generate_schema(user_requirements=user_requirements)
            # Save schema and requirements
            save_schema_to_python(BOMModel, schema_path)
            save_requirements(generator.item_requirements, BOMModel.__name__, requirements_path)
            requirements = generator.item_requirements
        # Extract data
        extractor = DataExtractor(config)
        result = extractor.extract(
            extraction_model=BOMModel,
            requirements=requirements,
            user_requirements=user_requirements,
            documents=[text],
        )
        if not result or not result[0]:
            raise ValueError(f"Failed to extract BOM data from {bom_file.filename}")
        # Convert to BOMData object
        bom_data = result[0]
        fee_flags = _extract_fee_flags_from_bom_text(text)
        bom = BOMData(
            material_id=bom_data.get("material_id", ""),
            type_designation=bom_data.get("type_designation", ""),
            dimensions=bom_data.get("dimensions", ""),
            material_grade=bom_data.get("material_grade", ""),
            cutting_required=fee_flags["cutting_required"] or bom_data.get("cutting_required", False),
            testing_required=fee_flags["testing_required"] or bom_data.get("testing_required", False),
            certificates_required=fee_flags["certificates_required"] or bom_data.get("certificates_required", False),
        )
        return bom
    finally:
        Path(tmp_path).unlink(missing_ok=True)
async def parse_pricing_file(pricing_file: UploadFile) -> list[PricingRow]:
    """Parse pricing Excel file with dynamic header detection.
    Supports both:
    - generic unit-price sheets
    - ABB sheets with Type/Conversion/Packing/Copper/KG Price/Weight columns
    """
    content_bytes = await pricing_file.read()
    if not content_bytes:
        return []
    def norm(value: object) -> str:
        return re.sub(r"[^a-z0-9]+", "", str(value).lower())
    def to_float(value: object) -> float:
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        raw = str(value).strip()
        if not raw or raw.lower() in {"nan", "none"} or "#" in raw:
            return 0.0
        return float(raw.replace(",", "."))
    try:
        workbook = load_workbook(io.BytesIO(content_bytes), data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        header_row_idx = None
        header_map: dict[str, int] = {}
        for idx, row in enumerate(rows):
            normalized = [norm(cell) for cell in row]
            row_map = {name: i for i, name in enumerate(normalized) if name}
            has_generic = any(k in row_map for k in ("type", "typedesignation", "partdesignation", "typepartdesignation")) and any(
                k in row_map for k in ("unitprice", "priceperunit", "price")
            )
            has_abb = "type" in row_map and "conversion" in row_map and "copper" in row_map
            if has_generic or has_abb:
                header_row_idx = idx
                header_map = row_map
                logger.info(f"Detected pricing header row at Excel row {idx + 1}")
                logger.info(f"Excel columns: {[str(v) for v in row]}")
                break
        if header_row_idx is None:
            logger.error("Could not find required columns in Excel file")
            return []
        def first_col(*names: str) -> int | None:
            for name in names:
                if name in header_map:
                    return header_map[name]
            return None
        item_col = first_col("itemno", "itemnumber", "materialid", "materialnumber", "id")
        type_col = first_col("type", "typedesignation", "partdesignation", "typepartdesignation")
        generic_price_col = first_col("unitprice", "priceperunit", "price")
        abb_price_col = first_col("kgpriceexclmachining", "kgprice", "total")
        cutting_col = first_col("cuttingfee", "cutting")
        testing_col = first_col("testingfee", "testing")
        cert_col = first_col("certificatefee", "certfee", "certificationfee")
        logger.info(
            "Matched columns - "
            f"Item: {item_col}, Type: {type_col}, Generic Price: {generic_price_col}, ABB Price: {abb_price_col}, "
            f"Cutting: {cutting_col}, Testing: {testing_col}, Cert: {cert_col}"
        )
        pricing_rows: list[PricingRow] = []
        for row in rows[header_row_idx + 1 :]:
            if type_col is None or type_col >= len(row):
                continue
            material_id = (
                str(row[item_col] or "").strip()
                if item_col is not None and item_col < len(row)
                else ""
            )
            type_designation = str(row[type_col] or "").strip()
            if not type_designation or type_designation.lower() in {"nan", "none"}:
                continue
            try:
                if generic_price_col is not None and generic_price_col < len(row):
                    unit_price = to_float(row[generic_price_col])
                    cutting_fee = to_float(row[cutting_col]) if cutting_col is not None and cutting_col < len(row) else 0.0
                    testing_fee = to_float(row[testing_col]) if testing_col is not None and testing_col < len(row) else 0.0
                    cert_fee = to_float(row[cert_col]) if cert_col is not None and cert_col < len(row) else 0.0
                else:
                    unit_price = to_float(row[abb_price_col]) if abb_price_col is not None and abb_price_col < len(row) else 0.0
                    cutting_fee = 0.0
                    testing_fee = 0.0
                    cert_fee = 0.0
                if unit_price == 0:
                    continue
                pricing_rows.append(
                    PricingRow(
                        material_id=material_id or None,
                        type_designation=type_designation,
                        unit_price=unit_price,
                        cutting_fee=cutting_fee,
                        testing_fee=testing_fee,
                        cert_fee=cert_fee,
                    )
                )
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping row due to parsing error: {e}")
                continue
        logger.info(f"Parsed {len(pricing_rows)} pricing rows from {pricing_file.filename}")
        return pricing_rows
    except Exception as e:
        logger.error(f"Error parsing pricing file: {e}", exc_info=True)
        return []
def generate_pdf(
    po_number: str,
    customer: str,
    items: list[EnrichedItem],
    summary: OrderSummary,
    customer_address: str | None = None,
    delivery_address: str | None = None,
    invoicing_address: str | None = None,
) -> bytes:
    """Generate a structured order-draft PDF with demo-specific branding."""
    logo_path = Path(__file__).parent.parent.parent / "public" / "logo.png"
    def money(value: float | None) -> str:
        if value is None:
            return "-"
        return f"USD {value:,.2f}"
    def split_text(pdf: FPDF, value: str, width: float) -> list[str]:
        text_value = str(value or "").replace("\r", "")
        paragraphs = text_value.split("\n")
        lines: list[str] = []
        for paragraph in paragraphs:
            words = paragraph.split()
            if not words:
                lines.append("")
                continue
            current = words[0]
            for word in words[1:]:
                candidate = f"{current} {word}"
                if pdf.get_string_width(candidate) <= width:
                    current = candidate
                else:
                    lines.append(current)
                    current = word
            lines.append(current)
        return lines or [""]
    def ensure_space(pdf: FPDF, height: float) -> None:
        if pdf.get_y() + height > pdf.page_break_trigger:
            pdf.add_page()
    def render_info_block(pdf: FPDF, x: float, y: float, width: float, title: str, body: str) -> float:
        pdf.set_xy(x, y)
        pdf.set_fill_color(245, 247, 250)
        pdf.set_draw_color(220, 225, 232)
        pdf.set_line_width(0.2)
        body_font_size = 9
        line_height = 4.3
        inner_width = width - 6
        body_lines: list[str] = []
        pdf.set_font("Helvetica", "", body_font_size)
        for paragraph in body.split("\n"):
            body_lines.extend(split_text(pdf, paragraph, inner_width))
        body_lines = body_lines or [""]
        box_height = 8 + len(body_lines) * line_height + 4
        pdf.rect(x, y, width, box_height, style="FD")
        pdf.set_xy(x + 3, y + 5)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(60, 70, 85)
        pdf.cell(inner_width, 0, title)
        pdf.set_xy(x + 3, y + 10)
        pdf.set_font("Helvetica", "", body_font_size)
        pdf.set_text_color(15, 23, 42)
        for line in body_lines:
            pdf.cell(inner_width, line_height, line, new_x="LMARGIN", new_y="NEXT")
            pdf.set_x(x + 3)
        return box_height
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()
    page_left = 12
    page_right = 198
    # Header
    if logo_path.exists():
        pdf.image(str(logo_path), x=page_left, y=12, w=48)
    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_xy(120, 15)
    pdf.cell(68, 9, "Order Draft", align="R")
    pdf.set_draw_color(210, 214, 220)
    pdf.set_line_width(0.4)
    pdf.line(page_left, 28, page_right, 28)
    left_x = page_left
    right_x = 125
    top_y = 34
    left_width = 104
    right_width = page_right - right_x
    buyer_body = customer or "Not available"
    delivery_body = delivery_address or customer_address or customer or "Not available"
    invoice_body = invoicing_address or customer_address or customer or "Not available"
    left_y = top_y
    left_y += render_info_block(pdf, left_x, left_y, left_width, "Buyer", buyer_body) + 4
    left_y += render_info_block(pdf, left_x, left_y, left_width, "Delivery address", delivery_body) + 4
    left_y += render_info_block(pdf, left_x, left_y, left_width, "Invoicing address", invoice_body)
    order_details = [
        ("Date", datetime.now().strftime("%d/%m/%Y")),
        ("Order draft no.", f"DRAFT-{po_number or 'N/A'}"),
        ("Your reference", po_number or "-"),
        ("Currency", "USD"),
        ("Total lines", str(summary.total_items)),
        ("Total quantity", f"{summary.total_quantity}"),
    ]
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(60, 70, 85)
    details_height = 8 + len(order_details) * 6 + 4
    pdf.set_fill_color(245, 247, 250)
    pdf.set_draw_color(220, 225, 232)
    pdf.rect(right_x, top_y, right_width, details_height, style="FD")
    pdf.set_xy(right_x + 3, top_y + 5)
    pdf.cell(right_width - 6, 0, "Order details")
    detail_y = top_y + 12
    for label, value in order_details:
        pdf.set_xy(right_x + 3, detail_y)
        pdf.set_font("Helvetica", "B", 8.5)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(28, 4.5, label)
        pdf.set_font("Helvetica", "", 8.5)
        pdf.set_text_color(15, 23, 42)
        pdf.multi_cell(right_width - 35, 4.5, value)
        detail_y = pdf.get_y() + 1
    y = max(left_y, top_y + details_height) + 8
    ensure_space(pdf, 18)
    pdf.set_xy(page_left, y)
    pdf.set_font("Helvetica", "B", 8.5)
    pdf.set_fill_color(236, 240, 244)
    pdf.set_text_color(20, 20, 20)
    col_widths = [20, 67, 12, 22, 28, 37]
    headers = ["Material", "Description", "Qty", "Unit Price", "Fees", "Line Total"]
    for width, label in zip(col_widths, headers):
        pdf.cell(width, 8, label, border=1, align="C", fill=True)
    pdf.ln(8)
    for item in items:
        pdf.set_font("Helvetica", "", 8)
        description_lines = split_text(
            pdf,
            "\n".join(
                [
                    item.description,
                    f"Type: {item.type_designation}" if item.type_designation else "",
                    f"Delivery: {item.delivery_date}" if item.delivery_date else "",
                    f"Error: {item.error}" if item.error else "",
                ]
            ).strip(),
            col_widths[1] - 4,
        )
        fee_text = "\n".join(
            [
                f"Cut {item.cutting_fee:.2f}",
                f"Test {item.testing_fee:.2f}",
                f"Cert {item.cert_fee:.2f}",
            ]
        )
        fee_lines = split_text(pdf, fee_text, col_widths[4] - 4)
        row_line_count = max(1, len(description_lines), len(fee_lines))
        row_height = max(8, row_line_count * 4.3 + 2)
        ensure_space(pdf, row_height)
        x = page_left
        y = pdf.get_y()
        cells = [
            item.material,
            description_lines,
            str(item.quantity),
            money(item.unit_price),
            fee_lines,
            money(item.line_total),
        ]
        for idx, width in enumerate(col_widths):
            pdf.rect(x, y, width, row_height)
            pdf.set_xy(x + 2, y + 2)
            if idx == 1:
                for line in description_lines:
                    pdf.cell(width - 4, 4.2, line, new_x="LMARGIN", new_y="NEXT")
                    pdf.set_x(x + 2)
            elif idx == 4:
                pdf.set_font("Helvetica", "", 7.5)
                for line in fee_lines:
                    pdf.cell(width - 4, 4.0, line, new_x="LMARGIN", new_y="NEXT")
                    pdf.set_x(x + 2)
            else:
                align = "R" if idx in {2, 3, 5} else "L"
                pdf.set_font("Helvetica", "", 8)
                pdf.cell(width - 4, 4.5, str(cells[idx]), align=align)
            x += width
        pdf.set_y(y + row_height)
    y = pdf.get_y() + 8
    ensure_space(pdf, 30)
    summary_x = 118
    summary_w = page_right - summary_x
    pdf.set_fill_color(245, 247, 250)
    pdf.set_draw_color(220, 225, 232)
    pdf.rect(summary_x, y, summary_w, 30, style="FD")
    pdf.set_xy(summary_x + 3, y + 5)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(20, 20, 20)
    pdf.cell(summary_w - 6, 0, "Order summary")
    summary_rows = [
        ("Material subtotal", money(summary.material_subtotal)),
        ("Processing fees", money(summary.total_fees)),
        ("Grand total", money(summary.grand_total)),
    ]
    row_y = y + 12
    for label, value in summary_rows:
        pdf.set_xy(summary_x + 3, row_y)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(70, 70, 70)
        pdf.cell(40, 5, label)
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(summary_w - 46, 5, value, align="R")
        row_y += 6.5
    output = pdf.output(dest="S")
    if isinstance(output, str):
        return output.encode("latin1")
    return bytes(output)
# ============================================================================
# API Endpoints
# ============================================================================
@router.post("/process")
async def process_order(
    po_file: Annotated[UploadFile, File(description="Purchase Order PDF")],
    bom_files: Annotated[list[UploadFile], File(description="BOM PDF files")],
    pricing_file: Annotated[UploadFile, File(description="Pricing CSV/Excel file")],
) -> StreamingResponse:
    """
    Process Luvata ABB order with streaming progress updates.
    Steps:
    1. Extract PO data from PDF
    2. Extract BOM data from PDFs
    3. Parse pricing table
    4. Match and calculate pricing
    5. Return enriched order data
    """
    async def generate():
        try:
            # Validate inputs
            if not po_file:
                yield sse_event("error", {"message": "Purchase order file required"})
                return
            if not bom_files or len(bom_files) == 0:
                yield sse_event("error", {"message": "At least one BOM file required"})
                return
            if not pricing_file:
                yield sse_event("error", {"message": "Pricing file required"})
                return
            yield sse_event("status", {"message": "Starting order processing..."})
            # 1. Extract PO data
            yield sse_event("status", {"message": "Extracting purchase order data..."})
            po = await extract_po_data(po_file)
            logger.info(f"Extracted PO: {po.po_number} with {len(po.items)} items")
            # 2. Extract BOM data
            yield sse_event("status", {"message": f"Extracting {len(bom_files)} BOMs..."})
            boms = []
            for bom_file in bom_files:
                bom_data = await extract_bom_data(bom_file)
                boms.append(bom_data)
                logger.info(f"Extracted BOM: {bom_data.material_id}")
            # 3. Parse pricing table
            yield sse_event("status", {"message": "Parsing pricing table..."})
            pricing_rows = await parse_pricing_file(pricing_file)
            logger.info(f"Parsed {len(pricing_rows)} pricing rows")
            # 4. Enrich products
            yield sse_event("status", {"message": "Calculating prices..."})
            enriched_items, errors = enrich_products(po, boms, pricing_rows)
            # 5. Calculate summary
            summary = calculate_summary(enriched_items)
            # 6. Generate PDF
            yield sse_event("status", {"message": "Generating PDF..."})
            job_id = str(uuid.uuid4())
            pdf_bytes = generate_pdf(
                po.po_number,
                po.customer,
                enriched_items,
                summary,
                customer_address=po.customer_address,
                delivery_address=po.delivery_address,
                invoicing_address=po.invoicing_address,
            )
            pdf_storage[job_id] = pdf_bytes
            logger.info(f"Generated PDF with job_id: {job_id}")
            # Build response
            response = ProcessOrderResponse(
                success=True,
                po_number=po.po_number,
                customer=po.customer,
                items=enriched_items,
                summary=summary,
                errors=errors,
                pdf_job_id=job_id,
            )
            yield sse_event("complete", response.model_dump())
        except Exception as e:
            logger.exception("Error processing order")
            yield sse_event("error", {"message": str(e)})
    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )
@router.get("/pdf/{job_id}")
async def download_pdf(job_id: str) -> Response:
    """Download generated order draft PDF"""
    pdf_bytes = pdf_storage.get(job_id)
    if not pdf_bytes:
        raise HTTPException(status_code=404, detail="PDF not found")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=order_draft_{job_id}.pdf"},
    )
