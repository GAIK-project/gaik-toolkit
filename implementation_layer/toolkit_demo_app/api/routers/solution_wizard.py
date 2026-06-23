"""Solution Configuration Wizard router.

Hosts an interactive GAIK Solution Configuration Wizard session per browser
client, driven by the Claude Agent SDK (`ClaudeSDKClient`) on Azure Foundry, and
streams the wizard's output to the frontend over Server-Sent Events.

Design (mirrors api/routers/rag.py session pattern + the CLI runner
implementation_layer/solution_wizard/run_wizard_interactive.py):

  * One `ClaudeSDKClient` per session, kept alive in WIZARD_SESSIONS so the
    conversation context persists across user messages.
  * Each user message = one SSE request that streams exactly one wizard turn and
    ends at that turn's ResultMessage. Turn-end is NOT session-end -- the client
    stays connected for the next message.
  * The wizard runs with cwd = the solution_wizard package dir (so its
    `python scripts/...` calls resolve) and writes all artifacts into a
    per-session workspace the user can list/download.
"""

from __future__ import annotations

import asyncio
import base64
import csv
import logging
import os
import shutil
import tempfile
import time
import uuid
from collections.abc import AsyncGenerator
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

try:
    from utils import sse_event
except ImportError:
    from api.utils import sse_event

from claude_agent_sdk import (
    AssistantMessage,
    ClaudeAgentOptions,
    ClaudeSDKClient,
    ResultMessage,
    TextBlock,
    ToolUseBlock,
)

# StreamEvent (token-level partials) is exported from the types submodule, not
# always the top level. Optional across SDK versions.
try:
    from claude_agent_sdk.types import StreamEvent  # type: ignore
except ImportError:  # pragma: no cover
    try:
        from claude_agent_sdk import StreamEvent  # type: ignore
    except ImportError:
        StreamEvent = None  # type: ignore

router = APIRouter()
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths & config
# ---------------------------------------------------------------------------

# Wizard assets live at <repo>/implementation_layer/solution_wizard. In the
# source tree that is parents[4] of this file; in the container image the api/
# directory is flattened to /app and the wizard assets are copied to
# /app/implementation_layer/solution_wizard (parents[1]). Resolve defensively —
# an unguarded parents[4] raises IndexError in the container and would disable
# the whole router at import time.
_HERE = Path(__file__).resolve()


def _resolve_wizard_dir() -> Path | None:
    env_dir = os.getenv("WIZARD_DIR", "").strip()
    if env_dir:
        return Path(env_dir)
    candidates = [
        parent / "implementation_layer" / "solution_wizard"
        for parent in (list(_HERE.parents[4:5]) + [_HERE.parents[1]])
    ]
    return next((c for c in candidates if c.is_dir()), None)


WIZARD_DIR = _resolve_wizard_dir()
WORKSPACES_DIR = Path(
    os.getenv("WIZARD_WORKSPACES_DIR", "").strip() or _HERE.parents[1] / ".wizard_workspaces"
)

REQUIRED_FOUNDRY_VARS = [
    "CLAUDE_CODE_USE_FOUNDRY",
    "ANTHROPIC_FOUNDRY_API_KEY",
    "ANTHROPIC_FOUNDRY_RESOURCE",
]
DEFAULT_MODEL = "claude-sonnet-4-6"
SESSION_IDLE_SECONDS = 30 * 60  # reap sessions idle longer than this

# ---------------------------------------------------------------------------
# Session store
# ---------------------------------------------------------------------------

# session_id -> {client, output_dir, lock, last_active}
WIZARD_SESSIONS: dict[str, dict] = {}


class FileAttachment(BaseModel):
    name: str
    mime_type: str
    data: str  # base64 data URL: "data:<mime>;base64,<bytes>"


class MessageRequest(BaseModel):
    text: str
    files: list[FileAttachment] = []


def _parse_pdf_attachment(file_path: str, original_name: str) -> str:
    """Parse a PDF attachment, including OCR-capable fallback for scanned PDFs."""
    from gaik.software_components.parsers import PyMuPDFParser

    result = PyMuPDFParser().parse_document(file_path)
    text = (result.get("text_content") or "").strip()
    if text:
        logger.info(
            "Parsed wizard attachment %s via PyMuPDF (%d chars)",
            original_name,
            len(text),
        )
        return text

    api_base = os.getenv("DOCLING_API_BASE") or os.getenv("API_BASE")
    password = os.getenv("DOCLING_API_PASSWORD") or os.getenv("PASSWORD")
    if not api_base or not password:
        logger.warning(
            "PyMuPDF extracted no text from wizard attachment %s and Docling API is not configured",
            original_name,
        )
        return ""

    try:
        from gaik.software_components.parsers.docling_api_client import DoclingApiClientParser

        parser = DoclingApiClientParser(api_base=api_base, password=password)
        docling_result = parser.parse_document(file_path)
        markdown = (
            docling_result.get("parsed_markdown") or docling_result.get("text_content") or ""
        ).strip()
        if markdown:
            logger.info(
                "Parsed wizard attachment %s via Docling API fallback (%d chars)",
                original_name,
                len(markdown),
            )
            return markdown
        logger.warning(
            "Docling API returned empty markdown for wizard attachment %s", original_name
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "Docling API fallback failed for wizard attachment %s: %s",
            original_name,
            exc,
        )
    return ""


def _extract_text_from_attachment(attachment: FileAttachment) -> str:
    """Extract plain text from an uploaded file using GAIK parsers for PDF/DOCX."""
    raw = attachment.data
    if "," in raw:
        raw = raw.split(",", 1)[1]
    file_bytes = base64.b64decode(raw)

    ext = Path(attachment.name).suffix.lower()

    if ext in (".txt", ".md"):
        return file_bytes.decode("utf-8", errors="replace")

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        if ext == ".pdf":
            return _parse_pdf_attachment(tmp_path, attachment.name)
        elif ext in (".docx", ".doc"):
            from gaik.software_components.parsers import DocxParser

            result = DocxParser().parse_document(tmp_path)
            return result.get("text_content", "")
        elif ext == ".csv":
            with open(tmp_path, encoding="utf-8-sig", newline="") as fh:
                rows = list(csv.reader(fh))
            if not rows:
                return "_(empty CSV)_"
            header = "| " + " | ".join(str(c) for c in rows[0]) + " |"
            sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
            body = "\n".join("| " + " | ".join(str(c) for c in row) + " |" for row in rows[1:])
            return "\n".join([header, sep, body])
        elif ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook

            wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
            blocks = []
            for ws in wb.worksheets:
                rows = [
                    [("" if v is None else str(v)) for v in row]
                    for row in ws.iter_rows(values_only=True)
                ]
                rows = [r for r in rows if any(c.strip() for c in r)]
                if not rows:
                    continue
                header = "| " + " | ".join(rows[0]) + " |"
                sep = "| " + " | ".join("---" for _ in rows[0]) + " |"
                body = "\n".join("| " + " | ".join(r) + " |" for r in rows[1:])
                blocks.append(f"### Sheet: {ws.title}\n\n" + "\n".join([header, sep, body]))
            wb.close()
            return "\n\n".join(blocks) if blocks else "_(no tabular data found)_"
        elif ext in (".jpg", ".jpeg", ".png", ".webp", ".tiff", ".gif"):
            from gaik.software_components.config import get_openai_config
            from gaik.software_components.parsers import VisionParser

            config = get_openai_config(use_azure=bool(os.getenv("AZURE_API_KEY")))
            parser = VisionParser(openai_config=config)
            return parser.convert_image(tmp_path)
        elif ext in (".mp3", ".mp4", ".wav", ".m4a", ".ogg", ".webm", ".flac", ".mpeg", ".mpga"):
            try:
                from utils import get_api_config
            except ImportError:
                from api.utils import get_api_config
            from gaik.software_components.transcriber import Transcriber

            transcriber_workspace = tempfile.mkdtemp()
            try:
                transcriber = Transcriber(
                    api_config=get_api_config(),
                    output_dir=transcriber_workspace,
                )
                result = transcriber.transcribe(file_path=tmp_path)
                return result.raw_transcript
            finally:
                shutil.rmtree(transcriber_workspace, ignore_errors=True)
        else:
            return file_bytes.decode("utf-8", errors="replace")
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _foundry_env() -> dict[str, str]:
    """Full subprocess env (inherit + Foundry vars). Raises if creds missing."""
    missing = [k for k in REQUIRED_FOUNDRY_VARS if not (os.getenv(k) or "").strip()]
    if missing:
        raise HTTPException(
            status_code=500,
            detail=(
                "Solution Wizard is not configured: missing "
                + ", ".join(missing)
                + ". Add the Foundry variables to .env.local."
            ),
        )
    env = dict(os.environ)
    env.setdefault("API_TIMEOUT_MS", "600000")
    env.setdefault("DISABLE_TELEMETRY", "1")
    return env


def _model() -> str:
    return (
        os.getenv("ANTHROPIC_MODEL") or os.getenv("ANTHROPIC_DEFAULT_SONNET_MODEL") or DEFAULT_MODEL
    ).strip()


def _build_options(output_dir: Path) -> ClaudeAgentOptions:
    return ClaudeAgentOptions(
        cwd=str(WIZARD_DIR),
        add_dirs=[str(output_dir)],
        env=_foundry_env(),
        model=_model(),
        allowed_tools=["Read", "Grep", "Glob", "Write", "Edit", "Bash"],
        permission_mode="bypassPermissions",
        setting_sources=[],
        include_partial_messages=True,  # token-level streaming
    )


def _bootstrap_prompt(output_dir: Path) -> str:
    """Internal first message: invoke the skill, pre-set the output dir, start."""
    return f"""\
/solution-wizard

You are running an interactive GAIK Solution Configuration Wizard session in a
web chat. Load and follow the complete wizard instructions from ./SKILL.md (you
are already in the wizard directory).

The output directory is pre-selected and managed by the server:
{output_dir}
Use it directly and write ALL generated files there. Do NOT ask the user where
to save files, and never write anywhere else.

IMPORTANT INSTRUCTIONS FOR THIS WEB SESSION:
- The web UI has already displayed a welcome message. Do NOT greet the user or
  introduce yourself. Start the conversation directly.
- Never mention the output directory path to the user. File management is handled
  invisibly by the server.
- The user's FIRST message will be their use-case description (Step 1.2 of
  Phase 1). Acknowledge it briefly (1-2 sentences: pattern classification +
  what you understood), then move straight into Phase 2 requirement collection.
- Follow Phase 2 in full before moving to component selection:
    Round 1 — business context (current process, pain points, intended users,
              reviewers, stakeholders, success criteria, expected value, risks,
              PoC goal). Carry over anything already stated; only ask what is
              genuinely still missing.
    Round 2 — technical spec (language/locale, model provider preference,
              security constraints, integration targets, evaluation needs,
              runtime interface).
    Round 3 — target output fields (for extraction use cases only).
  Do NOT jump to component selection or field design before completing Rounds 1
  and 2 with the user.
- Ask one or two questions per message and wait for the reply. Use Markdown.
"""


# ---------------------------------------------------------------------------
# Streaming bridge
# ---------------------------------------------------------------------------


def _tool_summary(block: ToolUseBlock) -> str:
    """Short human-readable activity line for a tool use."""
    inp = block.input if isinstance(block.input, dict) else {}
    if block.name == "Bash":
        return str(inp.get("command", ""))[:160]
    for key in ("file_path", "path", "pattern", "command"):
        if key in inp:
            return f"{inp[key]}"[:160]
    return ""


_HEARTBEAT_INTERVAL = 20  # seconds between SSE keep-alive pings


async def _receive_with_heartbeat(client: ClaudeSDKClient):
    """Async generator that interleaves ``receive_response()`` messages with
    periodic ``None`` heartbeat sentinels.

    When the wizard is running a heavy tool call (generate_bpmn, scaffold_poc,
    etc.) it can be silent for 30–60 s. Without periodic data, browsers, Bun's
    HTTP layer, and upstream proxies treat the SSE connection as idle and close
    it. Yielding a sentinel every ``_HEARTBEAT_INTERVAL`` seconds lets the
    caller emit an SSE comment (ignored by the frontend) that keeps the TCP
    connection alive.
    """
    queue: asyncio.Queue = asyncio.Queue()

    async def _drain() -> None:
        async for msg in client.receive_response():
            await queue.put(msg)
        await queue.put(StopAsyncIteration)  # sentinel: stream finished

    async def _ping() -> None:
        while True:
            await asyncio.sleep(_HEARTBEAT_INTERVAL)
            await queue.put(None)  # heartbeat sentinel

    drain_task = asyncio.create_task(_drain())
    ping_task = asyncio.create_task(_ping())
    try:
        while True:
            item = await queue.get()
            if item is StopAsyncIteration:
                break
            yield item  # None = heartbeat, anything else = real message
    finally:
        ping_task.cancel()
        drain_task.cancel()


async def _stream_turn(session: dict, *, _silent_retries: int = 0) -> AsyncGenerator[str, None]:
    """Stream one wizard turn as SSE. Ends at the turn's ResultMessage.

    * Emits an SSE comment (keep-alive ping) every 20 s during silent
      stretches so the connection is never dropped by an idle-connection
      timeout in the browser, Bun, or an upstream proxy.
    * If a turn ends with no visible text (Claude did only tool calls or paused
      before generating its reply), transparently sends a quiet "Please
      continue." nudge and loops for one more turn — up to two retries.
    """
    client: ClaudeSDKClient = session["client"]
    had_text = False
    try:
        async for message in _receive_with_heartbeat(client):
            if message is None:
                # Heartbeat — emit an SSE comment so the connection stays alive.
                yield ": keep-alive\n\n"
                continue

            # Token-level partial deltas (preferred).
            if StreamEvent is not None and isinstance(message, StreamEvent):
                delta = _extract_stream_text(message)
                if delta:
                    had_text = True
                    yield sse_event("text_delta", {"text": delta})
                else:
                    thinking = _extract_stream_thinking(message)
                    if thinking:
                        yield sse_event("thinking_delta", {"text": thinking})
                continue

            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock):
                        # When partials are on, full TextBlocks duplicate the
                        # deltas -- skip them to avoid double-rendering.
                        if StreamEvent is None:
                            had_text = True
                            yield sse_event("text_delta", {"text": block.text})
                    elif isinstance(block, ToolUseBlock):
                        yield sse_event(
                            "tool_use",
                            {"name": block.name, "summary": _tool_summary(block)},
                        )
            elif isinstance(message, ResultMessage):
                if message.is_error:
                    yield sse_event(
                        "error",
                        {"message": message.result or f"stopped: {message.stop_reason}"},
                    )
                    break
                if not had_text and _silent_retries < 2:
                    # Turn ended silently — Claude paused before producing a
                    # visible response (common when a tool-heavy turn is followed
                    # by an implicit wait for input). Nudge it to continue
                    # without surfacing anything to the user.
                    await client.query("Please continue.")
                    async for chunk in _stream_turn(session, _silent_retries=_silent_retries + 1):
                        yield chunk
                else:
                    yield sse_event("done", {})
                break
    except Exception as exc:  # noqa: BLE001
        yield sse_event("error", {"message": str(exc)})
    finally:
        session["last_active"] = time.time()


def _extract_stream_text(event) -> str:
    """Pull incremental assistant text from a StreamEvent.

    `StreamEvent.event` is the raw Anthropic streaming event dict. We only want
    the assistant's visible text deltas, i.e. content_block_delta with a
    text_delta (NOT thinking_delta / input_json_delta for tool args).
    """
    ev = getattr(event, "event", None)
    if not isinstance(ev, dict):
        return ""
    if ev.get("type") != "content_block_delta":
        return ""
    delta = ev.get("delta") or {}
    if isinstance(delta, dict) and delta.get("type") == "text_delta":
        return delta.get("text", "") or ""
    return ""


def _extract_stream_thinking(event) -> str:
    """Pull incremental extended-thinking text from a StreamEvent, when the
    model emits thinking blocks. Surfaced to the UI as a collapsible
    "Reasoning" section the user can expand."""
    ev = getattr(event, "event", None)
    if not isinstance(ev, dict) or ev.get("type") != "content_block_delta":
        return ""
    delta = ev.get("delta") or {}
    if isinstance(delta, dict) and delta.get("type") == "thinking_delta":
        return delta.get("thinking", "") or ""
    return ""


def _sse_headers() -> dict:
    return {
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "X-Accel-Buffering": "no",
    }


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/start")
async def start_session() -> StreamingResponse:
    """Create a session and stream the wizard's opening turn.

    The first SSE event is `session {session_id}` so the client can store it.
    """
    if WIZARD_DIR is None or not (WIZARD_DIR / "SKILL.md").is_file():
        raise HTTPException(
            status_code=500,
            detail=(
                "Solution Wizard is not configured: wizard assets not found. "
                "Set WIZARD_DIR to the solution_wizard directory (must contain SKILL.md)."
            ),
        )
    _foundry_env()  # validate creds before creating anything

    session_id = uuid.uuid4().hex[:12]
    output_dir = WORKSPACES_DIR / session_id
    output_dir.mkdir(parents=True, exist_ok=True)

    client = ClaudeSDKClient(options=_build_options(output_dir))
    await client.connect()

    session = {
        "client": client,
        "output_dir": output_dir,
        "lock": asyncio.Lock(),
        "last_active": time.time(),
    }
    WIZARD_SESSIONS[session_id] = session

    async def gen() -> AsyncGenerator[str, None]:
        yield sse_event("session", {"session_id": session_id})
        async with session["lock"]:
            await client.query(_bootstrap_prompt(output_dir))
            async for chunk in _stream_turn(session):
                yield chunk

    return StreamingResponse(gen(), media_type="text/event-stream", headers=_sse_headers())


@router.post("/message/{session_id}")
async def send_message(session_id: str, body: MessageRequest) -> StreamingResponse:
    """Send a user message and stream the wizard's reply for this session."""
    session = WIZARD_SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found or expired.")
    if session["lock"].locked():
        raise HTTPException(status_code=409, detail="The wizard is still responding.")

    async def gen() -> AsyncGenerator[str, None]:
        async with session["lock"]:
            message_text = body.text
            if body.files:
                file_sections = []
                for f in body.files[:5]:
                    try:
                        content = _extract_text_from_attachment(f)
                    except Exception as exc:  # noqa: BLE001
                        content = f"[Could not extract text: {exc}]"
                    file_sections.append(
                        f'<attached_file name="{f.name}">\n{content}\n</attached_file>'
                    )
                message_text = "\n\n".join(file_sections) + "\n\n" + body.text
            await session["client"].query(message_text)
            async for chunk in _stream_turn(session):
                yield chunk

    return StreamingResponse(gen(), media_type="text/event-stream", headers=_sse_headers())


@router.get("/files/{session_id}")
async def list_files(session_id: str) -> dict:
    """List the artifacts generated in this session's workspace."""
    session = WIZARD_SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found or expired.")
    root: Path = session["output_dir"]
    files = [
        str(p.relative_to(root)).replace("\\", "/") for p in sorted(root.rglob("*")) if p.is_file()
    ]
    return {"files": files}


@router.get("/files/{session_id}/{file_path:path}")
async def download_file(session_id: str, file_path: str) -> FileResponse:
    """Download one artifact (path-traversal guarded to the workspace)."""
    session = WIZARD_SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found or expired.")
    root: Path = session["output_dir"].resolve()
    target = (root / file_path).resolve()
    if root not in target.parents and target != root:
        raise HTTPException(status_code=400, detail="Invalid path.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(target, filename=target.name)


@router.get("/download/{session_id}")
async def download_zip(session_id: str) -> FileResponse:
    """Download the whole session workspace (blueprint + diagrams + docs + PoC)
    as a single .zip archive."""
    session = WIZARD_SESSIONS.get(session_id)
    if session is None:
        raise HTTPException(status_code=404, detail="Session not found or expired.")
    root: Path = session["output_dir"]
    if not any(root.rglob("*")):
        raise HTTPException(status_code=404, detail="No files generated yet.")
    # Build the archive in a temp location (not inside the watched workspace).
    archive_base = Path(tempfile.gettempdir()) / f"gaik_wizard_{session_id}"
    zip_path = shutil.make_archive(str(archive_base), "zip", root_dir=str(root))
    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename="solution_package.zip",
    )


@router.delete("/end/{session_id}")
async def end_session(session_id: str) -> dict:
    """Disconnect and clean up a session."""
    session = WIZARD_SESSIONS.pop(session_id, None)
    if session is None:
        return {"status": "not_found"}
    try:
        await session["client"].disconnect()
    except Exception:  # noqa: BLE001
        pass
    shutil.rmtree(session["output_dir"], ignore_errors=True)
    return {"status": "ended"}


# ---------------------------------------------------------------------------
# Idle-session cleanup (registered in main.py lifespan)
# ---------------------------------------------------------------------------


async def cleanup_idle_sessions() -> None:
    """Background loop: disconnect + drop sessions idle beyond the timeout."""
    while True:
        await asyncio.sleep(600)
        cutoff = time.time() - SESSION_IDLE_SECONDS
        stale = [sid for sid, s in WIZARD_SESSIONS.items() if s["last_active"] < cutoff]
        for sid in stale:
            session = WIZARD_SESSIONS.pop(sid, None)
            if not session:
                continue
            try:
                await session["client"].disconnect()
            except Exception:  # noqa: BLE001
                pass
            shutil.rmtree(session["output_dir"], ignore_errors=True)
