"""Generate the sample files used by the tabular_agent examples and tests.

Everything here is synthetic -- invented company names and made-up numbers.
Nothing is derived from real or customer data, so the generated files are safe
to commit and to share.

Run:
    python make_fixtures.py

Creates, under ``input/``:
    sales_clean.csv        a tidy export: headers on row 1, no surprises
    sales_semicolon.csv    Nordic style: ';' delimiter, ',' decimals, UTF-8 BOM
    messy_report.xlsx      a human-facing report: title rows, merged-looking
                           headers, a blank spacer, subtotals and a notes block
    multi_sheet.xlsx       two related sheets, joinable on region
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

INPUT_DIR = Path(__file__).parent / "input"

REGIONS = ["North", "South", "East", "West"]

# (region, quarter, product, units, revenue_eur)
SALES: list[tuple[str, str, str, int, float]] = [
    ("North", "Q1", "Widget", 120, 14400.0),
    ("North", "Q2", "Widget", 135, 16200.0),
    ("North", "Q3", "Widget", 180, 21600.0),
    ("North", "Q1", "Gadget", 40, 12000.0),
    ("North", "Q2", "Gadget", 55, 16500.0),
    ("North", "Q3", "Gadget", 61, 18300.0),
    ("South", "Q1", "Widget", 90, 10800.0),
    ("South", "Q2", "Widget", 88, 10560.0),
    ("South", "Q3", "Widget", 94, 11280.0),
    ("South", "Q1", "Gadget", 30, 9000.0),
    ("South", "Q2", "Gadget", 28, 8400.0),
    ("South", "Q3", "Gadget", 35, 10500.0),
    ("East", "Q1", "Widget", 210, 25200.0),
    ("East", "Q2", "Widget", 245, 29400.0),
    ("East", "Q3", "Widget", 305, 36600.0),
    ("East", "Q1", "Gadget", 75, 22500.0),
    ("East", "Q2", "Gadget", 80, 24000.0),
    ("East", "Q3", "Gadget", 96, 28800.0),
    ("West", "Q1", "Widget", 60, 7200.0),
    ("West", "Q2", "Widget", 58, 6960.0),
    ("West", "Q3", "Widget", 51, 6120.0),
    ("West", "Q1", "Gadget", 20, 6000.0),
    ("West", "Q2", "Gadget", 22, 6600.0),
    ("West", "Q3", "Gadget", 19, 5700.0),
]

HEADCOUNT = [("North", 12), ("South", 9), ("East", 21), ("West", 6)]


def write_clean_csv(path: Path) -> None:
    """A tidy export -- the case the heuristics must handle with no LLM call."""
    lines = ["region,quarter,product,units,revenue_eur"]
    lines += [f"{r},{q},{p},{u},{rev:.2f}" for r, q, p, u, rev in SALES]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_semicolon_csv(path: Path) -> None:
    """Nordic conventions: ';' separator, ',' decimal mark, and a UTF-8 BOM."""
    lines = ["region;quarter;product;units;revenue_eur"]
    lines += [f"{r};{q};{p};{u};" + f"{rev:.2f}".replace(".", ",") for r, q, p, u, rev in SALES]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8-sig")


def write_messy_xlsx(path: Path) -> None:
    """A report built for human eyes, not for a parser.

    Row 0 is a title, row 1 a subtitle, row 2 blank, row 3 the real header.
    Subtotal rows are interleaved and a notes block trails the data -- exactly
    the shapes that make a naive "headers are on row 1" reader produce garbage.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales Report"

    sheet.append(["ACME Nordics -- Quarterly Sales Report"])
    sheet.append(["Generated for internal review -- figures in EUR"])
    sheet.append([])
    sheet.append(["Region", "Quarter", "Product", "Units", "Revenue (EUR)"])

    for region in REGIONS:
        rows = [s for s in SALES if s[0] == region]
        for _, quarter, product, units, revenue in rows:
            sheet.append([region, quarter, product, units, revenue])
        sheet.append(
            [f"{region} subtotal", None, None, sum(r[3] for r in rows), sum(r[4] for r in rows)]
        )

    sheet.append([])
    sheet.append(["Notes:"])
    sheet.append(["Figures are unaudited."])
    sheet.append(["Contact: reporting@example.invalid"])

    workbook.save(path)


def write_multi_sheet_xlsx(path: Path) -> None:
    """Two clean sheets that only answer a question when joined together."""
    workbook = Workbook()

    sales = workbook.active
    sales.title = "Sales"
    sales.append(["region", "quarter", "product", "units", "revenue_eur"])
    for row in SALES:
        sales.append(list(row))

    staff = workbook.create_sheet("Headcount")
    staff.append(["region", "employees"])
    for row in HEADCOUNT:
        staff.append(list(row))

    workbook.save(path)


def main() -> None:
    """Write every fixture into ``input/``."""
    INPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_clean_csv(INPUT_DIR / "sales_clean.csv")
    write_semicolon_csv(INPUT_DIR / "sales_semicolon.csv")
    write_messy_xlsx(INPUT_DIR / "messy_report.xlsx")
    write_multi_sheet_xlsx(INPUT_DIR / "multi_sheet.xlsx")
    for file in sorted(INPUT_DIR.iterdir()):
        print(f"wrote {file.relative_to(Path(__file__).parent)}")


if __name__ == "__main__":
    main()
