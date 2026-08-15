#!/usr/bin/env python3
"""Build the human-review workbook from results/comparison_data.json."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "results" / "comparison_data.json"

SHEETS = {
    "EQ1 Requirements": ("EQ1-", True),
    "EQ1 Diagnostics": ("EQ1-", False),
    "EQ2 Configuration": ("EQ2-", True),
    "EQ3 Package": ("EQ3-", True),
    "EQ4 Execution": ("EQ4-", True),
}
HEADERS = ["Check ID", "Parameter", "Oracle value", "Oracle source", "Run 1 generated value", "Run 1 evidence", "Run 1 verdict", "Run 2 generated value", "Run 2 evidence", "Run 2 verdict", "Evaluator notes"]
NAVY = "1F4E78"
PALE = "D9EAF7"
LIGHT = "F3F6F9"
GOLD = "FFF2CC"
THIN = Side(style="thin", color="B7C9DC")


def load():
    if not INPUT.exists():
        raise SystemExit("Missing results/comparison_data.json. Run collect_evidence.py --run all first.")
    return json.loads(INPUT.read_text(encoding="utf-8"))


def style_sheet(ws, title: str):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(2, ws.max_row)}"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    widths = [15, 28, 48, 30, 44, 58, 16, 44, 58, 16, 34]
    for index, width in enumerate(widths[:ws.max_column], start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        if len(row) >= 10:
            row[6].fill = PatternFill("solid", fgColor=GOLD)
            row[9].fill = PatternFill("solid", fgColor=GOLD)
    if ws.max_row >= 2 and ws.max_column >= 10:
        validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"G2:G{ws.max_row}")
        validation.add(f"J2:J{ws.max_row}")
        green = PatternFill("solid", fgColor="E2F0D9")
        red = PatternFill("solid", fgColor="F4CCCC")
        for col in ("G", "J"):
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator="equal", formula=['"Yes"'], fill=green))
            ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", CellIsRule(operator="equal", formula=['"No"'], fill=red))


def build_summary(wb, scenario: str, sheet_rows: dict[str, tuple[int, int]]):
    ws = wb.create_sheet("Summary", 1)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{scenario} evaluation summary"
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:H2")
    ws["A2"] = "Scores are populated by score_workbook.py after all required human verdicts are entered. EQ3 and EQ4 use an all-mandatory-checks pass rule per run."
    ws["A2"].font = Font(italic=True, color="44546A")
    ws["A2"].alignment = Alignment(wrap_text=True)
    for column, value in enumerate(["EQ", "Metric", "Run 1", "Run 2", "Combined score", "Completed", "Required verdicts", "Interpretation"], start=1):
        ws.cell(4, column, value)
    metrics = [
        ("EQ1", "Requirement capture recall", "Pooled row-level recall"),
        ("EQ2", "Configuration constraint satisfaction", "Pooled constraint satisfaction"),
        ("EQ3", "Valid solution package rate", "A run passes only if all EQ3 rows are Yes"),
        ("EQ4", "PoC execution success rate", "A run passes only if all EQ4 rows are Yes"),
    ]
    for row, values in enumerate(metrics, start=5):
        ws.cell(row, 1, values[0]); ws.cell(row, 2, values[1]); ws.cell(row, 8, values[2])
        ws.cell(row, 6, 0)
        key_sheet = {"EQ1": "EQ1 Requirements", "EQ2": "EQ2 Configuration", "EQ3": "EQ3 Package", "EQ4": "EQ4 Execution"}[values[0]]
        required = sheet_rows[key_sheet][1] - sheet_rows[key_sheet][0] + 1
        ws.cell(row, 7, required * 2)
    for column, value in enumerate(["Diagnostic", "Run 1", "Run 2", "Reporting", "", "", "", "Not included in headline scores"], start=1):
        ws.cell(10, column, value)
    ws["A11"] = "Unsupported assumptions"; ws["D11"] = "Yes/No diagnostic verdict counts"
    ws["A12"] = "Refinement attempts to success"; ws["D12"] = "0, 1, 2, 3, or N/A"; ws["H12"] = "Original EQ4 is unchanged"
    ws["A13"] = "Recovery status"; ws["D13"] = "successful_original, recovered, pending_refinement, or unsuccessful_after_maximum_refinements"; ws["H13"] = "Maximum three refinement attempts"
    ws["A14"] = "Note"; ws["H14"] = "Blank verdicts mean the evaluation is incomplete, not zero performance."
    for row in (4, 10):
        for cell in ws[row]:
            cell.fill = PatternFill("solid", fgColor=PALE if row == 4 else GOLD)
            cell.font = Font(bold=True)
            cell.border = Border(bottom=THIN)
    for row in ws.iter_rows(min_row=5, max_row=14, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
    widths = [15, 34, 17, 17, 20, 16, 20, 54]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in range(5, 9):
        for col in range(3, 6):
            ws.cell(row, col).number_format = "0.0%"


def main():
    data = load()
    scenario = data["scenario_id"]
    run1 = {item["check_id"]: item for item in data["runs"]["run_01"]["checks"]}
    run2 = {item["check_id"]: item for item in data["runs"]["run_02"]["checks"]}
    ordered = list(run1)
    wb = Workbook()
    wb.remove(wb.active)
    instructions = wb.create_sheet("Instructions")
    instructions.sheet_view.showGridLines = False
    instructions.merge_cells("A1:F1"); instructions["A1"] = f"{scenario} comparison workbook"
    instructions["A1"].fill = PatternFill("solid", fgColor=NAVY); instructions["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    instructions["A3"] = "Evaluator task"
    instructions["A4"] = "Review the oracle value, generated value, and evidence. Enter only Yes or No in the Run 1 and Run 2 verdict columns. Notes are optional. Do not edit generated evidence."
    instructions["A6"] = "Verdict rule"
    instructions["A7"] = "Yes means the generated result semantically satisfies the oracle. No means it is missing, contradictory, incomplete, invalid, or unsupported. NOT FOUND normally receives No."
    instructions["A9"] = "Scoring"
    instructions["A10"] = "EQ1 and EQ2 are row-level proportions. EQ3 and EQ4 pass a run only when every mandatory row is Yes. PoC recovery is an unscored diagnostic and never changes the baseline EQ4 result."
    instructions.column_dimensions["A"].width = 120
    for row in (3, 6, 9): instructions[f"A{row}"].font = Font(bold=True, size=13)
    for row in (4, 7, 10): instructions[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top"); instructions.row_dimensions[row].height = 42

    ranges = {}
    for sheet_name, (prefix, scored) in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        selected = [cid for cid in ordered if cid.startswith(prefix) and bool(run1[cid].get("scored", True)) is scored]
        for cid in selected:
            one, two = run1[cid], run2.get(cid, {})
            ws.append([cid, one["parameter"], one["oracle_value"], one.get("oracle_source", ""), one.get("generated_value", ""), one.get("evidence", ""), "", two.get("generated_value", ""), two.get("evidence", ""), "", ""])
        style_sheet(ws, sheet_name)
        ranges[sheet_name] = (2, max(1, ws.max_row))

    recovery = wb.create_sheet("PoC Recovery")
    recovery.append(["Field", "Run 1", "Run 2", "Interpretation"])
    rec1 = data["runs"]["run_01"].get("poc_recovery", {})
    rec2 = data["runs"]["run_02"].get("poc_recovery", {})
    fields = [
        ("Initial execution successful", "initial_execution_successful"),
        ("Refinement attempts to success", "refinement_attempts_to_success"),
        ("Final execution successful", "final_execution_successful"),
        ("Recovery status", "status"),
    ]
    for label, key in fields:
        recovery.append([label, rec1.get(key), rec2.get(key), "Unscored; EQ4 always uses attempt 0."])
    recovery.append(["Maximum refinement attempts", 3, 3, "Stop after first success; N/A after three failed refinements."])
    recovery.append(["EQ4 uses baseline only", True, True, "A refined success does not improve EQ4."])
    style_sheet(recovery, "PoC Recovery")
    recovery.column_dimensions["A"].width = 34; recovery.column_dimensions["B"].width = 24; recovery.column_dimensions["C"].width = 24; recovery.column_dimensions["D"].width = 62

    build_summary(wb, scenario, ranges)
    output = ROOT / "results" / f"{scenario}_comparison.xlsx"
    output.parent.mkdir(exist_ok=True)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(output)
    print(f"Workbook generated: {output}")


if __name__ == "__main__":
    main()
