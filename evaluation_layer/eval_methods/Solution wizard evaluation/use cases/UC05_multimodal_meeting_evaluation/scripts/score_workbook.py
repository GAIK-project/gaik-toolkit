#!/usr/bin/env python3
"""Score a completed comparison workbook and populate its Summary sheet."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
METRICS = {
    "EQ1": ("EQ1 Requirements", "row"),
    "EQ2": ("EQ2 Configuration", "row"),
    "EQ3": ("EQ3 Package", "all"),
    "EQ4": ("EQ4 Execution", "all"),
}


def verdicts(ws):
    headers = {cell.value: cell.column for cell in ws[1]}
    one, two, missing = [], [], []
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row, headers["Check ID"]).value:
            continue
        v1 = ws.cell(row, headers["Run 1 verdict"]).value
        v2 = ws.cell(row, headers["Run 2 verdict"]).value
        for run, value in (("run_01", v1), ("run_02", v2)):
            if value not in {"Yes", "No"}:
                missing.append({"sheet": ws.title, "row": row, "run": run})
        one.append(v1); two.append(v2)
    return one, two, missing


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", nargs="?")
    parser.add_argument("output", nargs="?")
    args = parser.parse_args()
    if args.workbook:
        workbook_path = Path(args.workbook)
    else:
        candidates = sorted((ROOT / "results").glob("*_comparison.xlsx"))
        if len(candidates) != 1:
            raise SystemExit("Specify the workbook path or keep exactly one *_comparison.xlsx in results/.")
        workbook_path = candidates[0]
    wb = load_workbook(workbook_path)
    scenario = workbook_path.stem.replace("_comparison", "")
    metrics, all_missing = {}, []
    for eq, (sheet_name, mode) in METRICS.items():
        one, two, missing = verdicts(wb[sheet_name])
        all_missing.extend(missing)
        complete = not missing and bool(one)
        if complete:
            if mode == "row":
                run1 = sum(v == "Yes" for v in one) / len(one)
                run2 = sum(v == "Yes" for v in two) / len(two)
                combined = (sum(v == "Yes" for v in one) + sum(v == "Yes" for v in two)) / (2 * len(one))
            else:
                run1 = int(all(v == "Yes" for v in one))
                run2 = int(all(v == "Yes" for v in two))
                combined = (run1 + run2) / 2
        else:
            run1 = run2 = combined = None
        metrics[eq] = {"complete": complete, "scored_checks_per_run": len(one), "run_01": run1, "run_02": run2, "combined": combined}

    recovery_ws = wb["PoC Recovery"]
    recovery_values = {str(recovery_ws.cell(row, 1).value): (recovery_ws.cell(row, 2).value, recovery_ws.cell(row, 3).value) for row in range(2, recovery_ws.max_row + 1)}
    recovery = {
        "scored_under_EQ4": False,
        "maximum_refinement_attempts": 3,
        "run_01": {
            "initial_execution_successful": recovery_values.get("Initial execution successful", (None, None))[0],
            "refinement_attempts_to_success": recovery_values.get("Refinement attempts to success", (None, None))[0],
            "final_execution_successful": recovery_values.get("Final execution successful", (None, None))[0],
            "status": recovery_values.get("Recovery status", (None, None))[0],
        },
        "run_02": {
            "initial_execution_successful": recovery_values.get("Initial execution successful", (None, None))[1],
            "refinement_attempts_to_success": recovery_values.get("Refinement attempts to success", (None, None))[1],
            "final_execution_successful": recovery_values.get("Final execution successful", (None, None))[1],
            "status": recovery_values.get("Recovery status", (None, None))[1],
        },
    }
    result = {"schema_version": "1.0.0", "scenario_id": scenario, "source_workbook": str(workbook_path.resolve()), "complete": not all_missing, "metrics": metrics, "diagnostics": {"poc_recovery": recovery}, "missing_verdicts": all_missing}
    output = Path(args.output) if args.output else ROOT / "results" / f"{scenario}_scores.json"
    output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    summary = wb["Summary"]
    for offset, eq in enumerate(("EQ1", "EQ2", "EQ3", "EQ4"), start=5):
        metric = metrics[eq]
        summary.cell(offset, 3, metric["run_01"])
        summary.cell(offset, 4, metric["run_02"])
        summary.cell(offset, 5, metric["combined"])
        completed = 0 if not metric["complete"] else metric["scored_checks_per_run"] * 2
        summary.cell(offset, 6, completed)
        summary.cell(offset, 7, metric["scored_checks_per_run"] * 2)
        for col in (3, 4, 5): summary.cell(offset, col).number_format = "0.0%"
    if "EQ1 Diagnostics" in wb.sheetnames:
        d1, d2, _ = verdicts(wb["EQ1 Diagnostics"])
        summary["B11"] = f"Yes: {sum(v == 'Yes' for v in d1)}; No: {sum(v == 'No' for v in d1)}"
        summary["C11"] = f"Yes: {sum(v == 'Yes' for v in d2)}; No: {sum(v == 'No' for v in d2)}"
    summary["B12"] = recovery["run_01"]["refinement_attempts_to_success"]
    summary["C12"] = recovery["run_02"]["refinement_attempts_to_success"]
    summary["B13"] = recovery["run_01"]["status"]
    summary["C13"] = recovery["run_02"]["status"]
    summary["B14"] = "Complete" if not all_missing else f"Incomplete: {len(all_missing)} verdicts missing"
    wb.save(workbook_path)
    print(f"Scores generated: {output}")
    print(f"Workbook Summary updated: {workbook_path}")


if __name__ == "__main__":
    main()
