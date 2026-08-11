#!/usr/bin/env python3
r"""Populate the UC03 comparison workbook using openpyxl.

Run from the package root:

    python .\scripts\build_workbook.py

Optional positional arguments:

    python .\scripts\build_workbook.py INPUT_JSON OUTPUT_XLSX
"""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError as error:  # pragma: no cover - depends on evaluator machine
    raise SystemExit(
        "This workbook builder requires openpyxl. Install it once with:\n"
        "  python -m pip install -r requirements-workbook.txt"
    ) from error


EXCEL_CELL_UTF16_LIMIT = 32767
TRUNCATION_MARKER = "\n[TRUNCATED FOR EXCEL CELL LIMIT]"

SHEET_GROUPS = {
    "EQ1 Requirements": "EQ1",
    "EQ1 Diagnostics": "EQ1_diagnostics",
    "EQ2 Configuration": "EQ2",
    "EQ3 Package": "EQ3",
    "EQ4 Execution": "EQ4",
}


def _utf16_units(text: str) -> int:
    return sum(2 if ord(character) > 0xFFFF else 1 for character in text)


def _truncate_for_excel(text: str) -> str:
    if _utf16_units(text) <= EXCEL_CELL_UTF16_LIMIT:
        return text
    available = EXCEL_CELL_UTF16_LIMIT - _utf16_units(TRUNCATION_MARKER)
    characters: list[str] = []
    used = 0
    for character in text:
        units = 2 if ord(character) > 0xFFFF else 1
        if used + units > available:
            break
        characters.append(character)
        used += units
    return "".join(characters) + TRUNCATION_MARKER


def _excel_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False)
    elif value is True:
        text = "Yes"
    elif value is False:
        text = "No"
    else:
        text = str(value)
    return _truncate_for_excel(ILLEGAL_CHARACTERS_RE.sub("", text))


def _display_value(value: Any) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return "" if value is None else str(value)


def _set_text(sheet: Any, coordinate: str, value: Any) -> None:
    cell = sheet[coordinate]
    text = _excel_text(value)
    if not text:
        cell.value = None
        return
    cell.value = text
    cell.data_type = "s"


def _run_results(
    comparison: dict[str, Any], run_name: str, group: str
) -> dict[str, dict[str, Any]]:
    rows = (
        comparison.get("runs", {})
        .get(run_name, {})
        .get("results", {})
        .get(group, [])
    )
    return {
        str(row.get("check_id")): row
        for row in rows
        if isinstance(row, dict) and row.get("check_id")
    }


def _template_rows(sheet: Any) -> dict[str, int]:
    rows: dict[str, int] = {}
    for row_number in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row_number, column=2).value
        if isinstance(value, str) and value.startswith("EQ"):
            rows[value] = row_number
    return rows


def _populate_evaluation_sheet(
    sheet: Any, group: str, comparison: dict[str, Any]
) -> None:
    run_01 = _run_results(comparison, "run_01", group)
    run_02 = _run_results(comparison, "run_02", group)
    check_ids = sorted(set(run_01) | set(run_02))
    template_rows = _template_rows(sheet)
    missing = [check_id for check_id in check_ids if check_id not in template_rows]
    if missing:
        raise KeyError(
            f"Checks absent from template sheet {sheet.title!r}: "
            + ", ".join(missing)
        )

    for check_id in check_ids:
        row_number = template_rows[check_id]
        first = run_01.get(check_id, {})
        second = run_02.get(check_id, {})
        values = {
            f"F{row_number}": first.get("generated_value", ""),
            f"G{row_number}": first.get("evidence", ""),
            f"H{row_number}": first.get("automatic_result", ""),
            f"I{row_number}": first.get("human_verdict", ""),
            f"J{row_number}": second.get("generated_value", ""),
            f"K{row_number}": second.get("evidence", ""),
            f"L{row_number}": second.get("automatic_result", ""),
            f"M{row_number}": second.get("human_verdict", ""),
            f"N{row_number}": " | ".join(
                str(note)
                for note in (
                    first.get("evaluator_notes", ""),
                    second.get("evaluator_notes", ""),
                )
                if note
            ),
        }
        for coordinate, value in values.items():
            _set_text(sheet, coordinate, value)


def _populate_recovery(sheet: Any, comparison: dict[str, Any]) -> None:
    run_01 = comparison.get("runs", {}).get("run_01", {}).get(
        "recovery_diagnostic", {}
    )
    run_02 = comparison.get("runs", {}).get("run_02", {}).get(
        "recovery_diagnostic", {}
    )
    for row_number, key in (
        (4, "initial_execution_successful"),
        (5, "refinement_attempts_to_success"),
        (6, "final_execution_successful"),
        (7, "status"),
    ):
        _set_text(sheet, f"B{row_number}", _display_value(run_01.get(key)))
        _set_text(sheet, f"C{row_number}", run_01.get("evidence", ""))
        _set_text(sheet, f"D{row_number}", _display_value(run_02.get(key)))
        _set_text(sheet, f"E{row_number}", run_02.get("evidence", ""))


def _validate_input(comparison: dict[str, Any], workbook: Any) -> None:
    if comparison.get("scenario_id") != "UC03":
        raise ValueError(
            "Expected comparison data for scenario UC03, received "
            f"{comparison.get('scenario_id')!r}."
        )
    for run_name in ("run_01", "run_02"):
        if run_name not in comparison.get("runs", {}):
            raise ValueError(f"Comparison data is missing {run_name}.")
    required_sheets = {*SHEET_GROUPS, "PoC Recovery"}
    missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
    if missing_sheets:
        raise ValueError(
            "Workbook template is missing sheets: " + ", ".join(missing_sheets)
        )


def build_workbook(
    comparison_path: Path, output_path: Path, template_path: Path
) -> None:
    if not comparison_path.is_file():
        raise FileNotFoundError(
            f"Comparison data not found: {comparison_path}. "
            "Run collect_evidence.py --run all first."
        )
    if not template_path.is_file():
        raise FileNotFoundError(f"Workbook template not found: {template_path}")

    comparison = json.loads(comparison_path.read_text(encoding="utf-8"))
    workbook = load_workbook(template_path, data_only=False)
    _validate_input(comparison, workbook)
    for sheet_name, group in SHEET_GROUPS.items():
        _populate_evaluation_sheet(workbook[sheet_name], group, comparison)
    _populate_recovery(workbook["PoC Recovery"], comparison)

    calculation = getattr(workbook, "calculation", None)
    if calculation is not None:
        calculation.fullCalcOnLoad = True
        calculation.forceFullCalc = True
        calculation.calcMode = "auto"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f"{output_path.stem}.", suffix=".xlsx", dir=output_path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        verification = load_workbook(temporary_path, read_only=True, data_only=False)
        verification.close()
        workbook.close()
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def main() -> int:
    package_root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "comparison",
        nargs="?",
        type=Path,
        default=package_root / "results" / "comparison_data.json",
    )
    parser.add_argument(
        "output",
        nargs="?",
        type=Path,
        default=package_root / "results" / "UC03_comparison.xlsx",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=package_root / "templates" / "UC03_comparison_template.xlsx",
    )
    arguments = parser.parse_args()
    build_workbook(
        arguments.comparison.resolve(),
        arguments.output.resolve(),
        arguments.template.resolve(),
    )
    print(f"Wrote {arguments.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
